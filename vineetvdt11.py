import pandas as pd
import os
import streamlit as st
import logging
from datetime import datetime
import openpyxl
import math
from math import radians, sin, cos, sqrt, atan2
import io
import base64
import difflib

# Configure logging
logging.basicConfig(filename='network_search.log', level=logging.ERROR,
                    format='%(asctime)s - %(levelname)s - %(message)s')

class NetworkSearchApp:
    def __init__(self):
        # Initialize data structures in session state
        if 'lte_data' not in st.session_state:
            st.session_state.lte_data = pd.DataFrame()
        if 'nr_data' not in st.session_state:
            st.session_state.nr_data = pd.DataFrame()
        if 'bbu_data' not in st.session_state:
            st.session_state.bbu_data = pd.DataFrame()
        if 'usid_index' not in st.session_state:
            st.session_state.usid_index = {}
        if 'bbu_index' not in st.session_state:
            st.session_state.bbu_index = {}
        if 'file_paths' not in st.session_state:
            st.session_state.file_paths = {}
        if 'matched_records' not in st.session_state:
            st.session_state.matched_records = []
        if 'points' not in st.session_state:
            st.session_state.points = []
        if 'master_point' not in st.session_state:
            st.session_state.master_point = None
        if 'vdt_data' not in st.session_state:
            st.session_state.vdt_data = pd.DataFrame(columns=["LTE Site", "NR Site"])
        if 'lte_tree_record_map' not in st.session_state:
            st.session_state.lte_tree_record_map = {}
        if 'nr_tree_record_map' not in st.session_state:
            st.session_state.nr_tree_record_map = {}
        if 'auto_generate' not in st.session_state:
            st.session_state.auto_generate = True
        if 'project_name' not in st.session_state:
            st.session_state.project_name = "ATT_STX_253"
        if 'lte_columns' not in st.session_state:
            st.session_state.lte_columns = ["Source", "Site", "cell", "CELLRANGE", "CRSGAIN", "QRXLEVMIN", "EARFCNDL"]
        if 'nr_columns' not in st.session_state:
            st.session_state.nr_columns = ["Source", "USID", "SITE", "NRCELL_NAME", "Digital Tilt", "Power", "PCI",
                                          "ADMINISTRATIVESTATE", "CELLBARRED", "CELLRESERVEDFOROPERATOR",
                                          "OPERATIONALSTATE", "CELLRANGE", "SSBFREQUENCY", "CONFIGURATION"]

        # Market mapping
        self.market_mapping = {
            "ATT_ARK1": "ATT_ARK_253",
            "ATT_NoCAL1": "ATT_NoCAL_253",
            "ATT_SoCAL1": "ATT_SoCAL_253",
            "ATT_STX": "ATT_STX_253"
        }

        # Enhanced mappings
        self.mappings = {
            "LTE": {
                "USID": ["REMOTE_USID", "CSS_USID", "USID"],
                "ENBID": ["ENBID"],
                "cell ID": ["CELLID"],
                "Site": ["MECONTEXT_ID", "SITE", "OSS_ENodeB"],
                "Azumuth": ["ATOLL_AZIMUTH", "AZIMUTH", "Atoll_AZIMUT", "Atoll_Az"],
                "Digital Tilt": ["DIGITALTILT", "DIGITAL_TILT", "TILT"],
                "cell": ["EUTRAN_CELL_FDD_ID", "CELL", "OSS_EUTRAN_CELL_FDD_ID", "CELL_NAME"],
                "height(Meter)": ["Atoll_HEIGHT_m", "HEIGHT", "ANTENNA_HEIGHT", "HEIGHT_M"],
                "PCI": ["PHYSICALLAYERCELLID", "PCI", "OSS_PCI", "Atoll_PCI"],
                "Power": ["CONFIGUREDMAXTXPOWER", "TX_POWER", "OSS_CONFIGUREDMAXTXPOWER"],
                "LATITUDE": ["LATITUDE", "LAT"],
                "LONGITUDE": ["LONGITUDE", "LON", "LONG"],
                "ADMINISTRATIVESTATE": ["ADMINISTRATIVE_STATE", "ADMIN_STATE"],
                "OPERATIONALSTATE": ["OPERATIONALSTATE", "OP_STATE"],
                "CELLRANGE": ["CELLRANGE"],
                "CRSGAIN": ["CRSGAIN"],
                "QRXLEVMIN": ["QRXLEVMIN"],
                "EARFCNDL": ["EARFCNDL"],
                "Electrical Tilt": ["ELECTRICAL_TILT", "Atoll_ET", "E_TILT"],
                "ED_Market": ["ED_MARKET", "ED_Market", "EDMARKET"]
            },
            "5GNR": {
                "USID": ["CSS_USID", "REMOTE_USID", "USID", "BBU_USID"],
                "NIC": ["NCI"],
                "gnb ID": ["GNBID", "GNB_ID"],
                "Site": ["CTS_COMMON_ID", "GNB_NAME", "GNODEB"],
                "Azumuth": ["Atoll_AZIMUT", "AZIMUTH", "ATOLL_AZIMUTH"],
                "Digital Tilt": ["DIGITALTILT", "DIGITAL_TILT", "USEDDIGITALTILT"],
                "cell": ["NRCELLDUID", "CELL", "NRCELL_NAME"],
                "height(Meter)": ["Atoll_HEIGHT_m", "HEIGHT", "ANTENNA_HEIGHT", "ATOLL_ANTENNA_HEIGHT"],
                "PCI": ["NRPCI", "PCI"],
                "Power": ["CONFIGUREDMAXTXPOWER", "TX_POWER", "POWER"],
                "LATITUDE": ["LAT", "LATITUDE"],
                "LONGITUDE": ["LONG", "LONGITUDE"],
                "ADMINISTRATIVESTATE": ["ADMINISTRATIVESTATE", "ADMIN_STATE"],
                "OPERATIONALSTATE": ["OPERATIONALSTATE", "OP_STATE"],
                "CELLBARRED": ["CELLBARRED"],
                "CELLRESERVEDFOROPERATOR": ["CELLRESERVEDFOROPERATOR"],
                "CELLRANGE": ["CELLRANGE"],
                "SSBFREQUENCY": ["SSBFREQUENCY"],
                "ARFCNDL": ["ARFCNDL", "ARFCN_DL", "DL_ARFCN", "NR_ARFCNDL"],
                "CONFIGURATION": ["CONFIGURATION"],
                "Electrical Tilt": ["ELECTRICAL_TILT", "Atoll_ET", "E_TILT"],
                "ED_Market": ["ED_MARKET", "ED_Market", "EDMARKET"],
                "BBU_TECH": ["BBU_TECH"],
                "GNB_SA_STATE": ["GNB_SA_STATE"],
                "CELL_SA_STATE": ["CELL_SA_STATE"],
                "CELL_TYPE": ["CELL_TYPE"],
                "ON_AIR": ["ON_AIR"],
                "DSS_LTECELL": ["DSS_LTECELL"],
                "NRTAC": ["NRTAC"]
            },
            "5GNR_BBU": {
                "USID": ["USID", "REMOTE_USID", "CSS_USID"],
                "NRCELL_NAME": ["NRCELLDUID", "CELL", "NRCELL_NAME"],
                "CONFIGURATION": ["CONFIGURATION"],
                "ED_Market": ["ED_MARKET", "ED_Market", "EDMARKET"]
            }
        }

        # Google Maps API Key (optional, can be set via environment or Secrets)
        self.api_key = os.environ.get("GOOGLE_MAPS_API_KEY", "")

        # Custom CSS for styling
        st.markdown(
            """
            <style>
            .stApp {
                background-color: #f0f4f8;
                color: #333;
            }
            .stButton>button {
                background-color: #1e90ff;
                color: white;
                border-radius: 5px;
                padding: 5px 15px;
                transition: all 0.3s;
            }
            .stButton>button:hover {
                background-color: #104e8b;
                transform: scale(1.05);
            }
            .stHeader {
                font-size: 24px;
                font-weight: bold;
                color: #1e90ff;
                margin-bottom: 10px;
            }
            @media (max-width: 600px) {
                .stContainer {
                    padding: 5px;
                }
                .stButton>button {
                    width: 100%;
                    margin-bottom: 5px;
                }
            }
            </style>
            """,
            unsafe_allow_html=True
        )

        # Create UI
        self.create_widgets()

    def update_status(self, message):
        """Update status with timestamp"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        st.session_state.status = f"[{timestamp}] {message}"
        st.write(st.session_state.status)

    def create_widgets(self):
        st.markdown("<h1 class='stHeader'>Network Data Search Tool</h1>", unsafe_allow_html=True)

        # Data Loading Section
        with st.container():
            st.subheader("Data Loading")
            uploaded_files = st.file_uploader("Upload Multiple Files", accept_multiple_files=True, type=["xlsx", "xls"])
            if uploaded_files:
                st.session_state.file_paths = {file.name: file for file in uploaded_files}
                st.write("Uploaded Files:", ", ".join(st.session_state.file_paths.keys()))
            if st.button("Load Selected Data"):
                self.load_data(uploaded_files)

        # Search Section
        with st.container():
            st.subheader("Search")
            col1, col2, col3 = st.columns([1, 2, 1])
            with col1:
                search_type = st.selectbox("Search By:", ["USID", "NIC", "gnb ID", "ENBID", "cell ID", "Site"], key="search_type")
            with col2:
                search_value = st.text_input("Value:", key="search_value")
            with col3:
                if st.button("Search"):
                    self.perform_search(search_type, search_value)

        # Tabs
        tabs = st.tabs(["Main Results", "LTE Parameters", "5G Parameters", "VDT Sheet", "Distance Calculator"])
        with tabs[0]:
            self.create_main_tab()
        with tabs[1]:
            self.create_lte_tab()
        with tabs[2]:
            self.create_5g_tab()
        with tabs[3]:
            self.create_vdt_tab()
        with tabs[4]:
            self.create_distance_tab()

    def create_main_tab(self):
        st.subheader("Main Results")
        columns = [
            "Source", "NIC", "gnb ID", "ENBID", "cell ID", "USID", "Site",
            "Azumuth", "Digital Tilt", "cell", "height(Meter)", "PCI", "Power",
            "LATITUDE", "LONGITUDE", "ADMINISTRATIVESTATE", "OPERATIONALSTATE"
        ]
        if st.session_state.matched_records:
            main_data = []
            for tech, record in st.session_state.matched_records:
                row = {"Source": tech}
                for key in self.mappings[tech]:
                    row[key] = self.get_column_value(record, self.mappings[tech][key], tech)
                main_data.append(row)
            df = pd.DataFrame(main_data, columns=columns)
            st.dataframe(df, use_container_width=True)
            col1, col2 = st.columns([1, 1])
            with col1:
                if st.button("Export to Excel"):
                    self.export_to_excel(df, "Main_Results.xlsx")
            with col2:
                if st.button("Clear Results"):
                    self.clear_results()
            if st.button("Use for Distance Calculation"):
                self.use_for_distance(df)

    def create_lte_tab(self):
        st.subheader("LTE Parameters")
        col1, col2, col3 = st.columns([1, 1, 1])
        with col1:
            if st.button("Add Column"):
                self.add_lte_column()
        with col2:
            lte_cr_type = st.selectbox("Generate CR for:", ["cellRange", "crsGain", "Electrical Tilt"], key="lte_cr_type")
        with col3:
            if st.button("Generate CR"):
                self.generate_lte_cr(lte_cr_type)
        if st.button("Export LTE Data to Excel"):
            self.export_lte_to_excel()
        if st.session_state.matched_records:
            lte_data = []
            for tech, record in st.session_state.matched_records:
                if tech == "LTE":
                    row = {
                        "Source": "LTE",
                        "Site": self.get_column_value(record, self.mappings["LTE"]["Site"], "LTE"),
                        "cell": self.get_column_value(record, self.mappings["LTE"]["cell"], "LTE"),
                        "CELLRANGE": self.get_column_value(record, self.mappings["LTE"]["CELLRANGE"], "LTE"),
                        "CRSGAIN": self.get_column_value(record, self.mappings["LTE"]["CRSGAIN"], "LTE"),
                        "QRXLEVMIN": self.get_column_value(record, self.mappings["LTE"]["QRXLEVMIN"], "LTE"),
                        "EARFCNDL": self.get_column_value(record, self.mappings["LTE"]["EARFCNDL"], "LTE")
                    }
                    for col in st.session_state.lte_columns:
                        if col not in row and col in self.mappings["LTE"]:
                            row[col] = self.get_column_value(record, self.mappings["LTE"][col], "LTE")
                    lte_data.append(row)
            if lte_data:
                df = pd.DataFrame(lte_data, columns=st.session_state.lte_columns)
                st.dataframe(df, use_container_width=True)
                if st.button("Export Selected to Excel"):
                    self.export_lte_selected_to_excel(df)
                if st.button("Copy with Headers"):
                    self.copy_with_headers(df)

    def create_5g_tab(self):
        st.subheader("5G Parameters")
        col1, col2, col3 = st.columns([1, 1, 1])
        with col1:
            if st.button("Add Column"):
                self.add_5g_column()
        with col2:
            nr_cr_type = st.selectbox("Generate CR for:", ["digitalTilt", "cellRange"], key="nr_cr_type")
        with col3:
            if st.button("Generate 5G CR"):
                self.generate_5g_cr(nr_cr_type)
        if st.button("Export 5G Data to Excel"):
            self.export_5g_to_excel()
        if st.session_state.matched_records:
            nr_data = []
            for tech, record in st.session_state.matched_records:
                if tech == "5GNR":
                    row = {
                        "Source": "5GNR",
                        "USID": self.get_column_value(record, self.mappings["5GNR"]["USID"], "5GNR"),
                        "SITE": self.get_column_value(record, self.mappings["5GNR"]["Site"], "5GNR"),
                        "NRCELL_NAME": self.get_column_value(record, self.mappings["5GNR"]["cell"], "5GNR"),
                        "Digital Tilt": self.get_column_value(record, self.mappings["5GNR"]["Digital Tilt"], "5GNR"),
                        "Power": self.get_column_value(record, self.mappings["5GNR"]["Power"], "5GNR"),
                        "PCI": self.get_column_value(record, self.mappings["5GNR"]["PCI"], "5GNR"),
                        "ADMINISTRATIVESTATE": self.get_column_value(record, self.mappings["5GNR"]["ADMINISTRATIVESTATE"], "5GNR"),
                        "CELLBARRED": self.get_column_value(record, self.mappings["5GNR"]["CELLBARRED"], "5GNR"),
                        "CELLRESERVEDFOROPERATOR": self.get_column_value(record, self.mappings["5GNR"]["CELLRESERVEDFOROPERATOR"], "5GNR"),
                        "OPERATIONALSTATE": self.get_column_value(record, self.mappings["5GNR"]["OPERATIONALSTATE"], "5GNR"),
                        "CELLRANGE": self.get_column_value(record, self.mappings["5GNR"]["CELLRANGE"], "5GNR"),
                        "SSBFREQUENCY": self.get_column_value(record, self.mappings["5GNR"]["SSBFREQUENCY"], "5GNR"),
                        "CONFIGURATION": self.get_column_value(record, self.mappings["5GNR"]["CONFIGURATION"], "5GNR")
                    }
                    for col in st.session_state.nr_columns:
                        if col not in row and col in self.mappings["5GNR"]:
                            row[col] = self.get_column_value(record, self.mappings["5GNR"][col], "5GNR")
                    nr_data.append(row)
            if nr_data:
                df = pd.DataFrame(nr_data, columns=st.session_state.nr_columns)
                st.dataframe(df, use_container_width=True)
                if st.button("Export Selected to Excel"):
                    self.export_5g_selected_to_excel(df)
                if st.button("Copy with Headers"):
                    self.copy_with_headers(df)

    def create_vdt_tab(self):
        st.subheader("VDT Sheet")
        col1, col2, col3 = st.columns([1, 1, 1])
        with col1:
            project_name = st.selectbox("Project Name:", list(self.market_mapping.values()), key="project_name")
        with col2:
            if st.button("Generate VDT Report"):
                self.generate_vdt_report(project_name)
        with col3:
            auto_generate = st.checkbox("Auto-Generate on Search", value=st.session_state.auto_generate, key="auto_generate")
        if not st.session_state.vdt_data.empty:
            st.dataframe(st.session_state.vdt_data, use_container_width=True)

    def create_distance_tab(self):
        st.subheader("Distance Calculator")
        col1, col2 = st.columns([1, 2])
        with col1:
            st.write("Add Point")
            point_name = st.text_input("Point Name:", key="point_name")
            lat = st.number_input("Latitude:", format="%.6f", key="lat")
            lon = st.number_input("Longitude:", format="%.6f", key="lon")
            if st.button("Add Point"):
                self.add_point(point_name, lat, lon)
            st.write("Add from Lat,Long")
            lat_long = st.text_input("Lat,Long (e.g., 12.34,78.56):", key="lat_long")
            lat_long_name = st.text_input("Point Name for Lat,Long:", key="lat_long_name")
            if st.button("Add Lat,Long"):
                self.add_lat_long(lat_long, lat_long_name)
            uploaded_points = st.file_uploader("Import Points from Excel", type=["xlsx", "xls"], key="points_upload")
            if uploaded_points:
                self.import_points_from_excel(uploaded_points)
            if st.session_state.points:
                points_df = pd.DataFrame(st.session_state.points, columns=["Name", "Latitude", "Longitude"])
                selected = st.multiselect("Select Points:", points_df["Name"].tolist(), key="points_select")
                col_a, col_b, col_c = st.columns([1, 1, 1])
                with col_a:
                    if st.button("Set as Master"):
                        if selected:
                            point = next(p for p in st.session_state.points if p[0] == selected[0])
                            self.set_master_point(point)
                with col_b:
                    if st.button("Remove Selected"):
                        self.remove_points(selected)
                with col_c:
                    if st.button("Clear All Points"):
                        self.clear_points()
        with col2:
            dist_tabs = st.tabs(["Results", "Map Visualization"])
            with dist_tabs[0]:
                if st.button("Calculate Path Distances"):
                    self.calculate_path_distances()
                if st.button("Calculate from Master"):
                    self.calculate_from_master()
                if 'distance_results' in st.session_state and st.session_state.distance_results:
                    st.text_area("Distance Results:", st.session_state.distance_results, height=300, key="distance_results")
            with dist_tabs[1]:
                map_type = st.selectbox("Map Type:", ["roadmap", "satellite", "hybrid", "terrain"], key="map_type")
                map_zoom = st.number_input("Zoom:", min_value=1, max_value=20, value=12, key="map_zoom")
                if st.button("Show Map"):
                    self.show_map(map_type, map_zoom)
                if st.session_state.points:
                    st.write("Map Legend:")
                    for i, (name, _, _) in enumerate(st.session_state.points):
                        st.write(f"{chr(65 + i)}: {name}")
                    if st.session_state.master_point:
                        st.write(f"Master Point: {st.session_state.master_point[0]} (blue)")

    def load_data(self, files):
        try:
            new_lte_data = []
            new_nr_data = []
            new_bbu_data = []
            for file in files:
                df = pd.read_excel(file)
                file_name = file.name.lower()
                if "lte" in file_name:
                    new_lte_data.append(df)
                elif "nr" in file_name:
                    new_nr_data.append(df)
                elif "bbu" in file_name:
                    new_bbu_data.append(df)
            st.session_state.lte_data = pd.concat(new_lte_data, ignore_index=True) if new_lte_data else pd.DataFrame()
            st.session_state.nr_data = pd.concat(new_nr_data, ignore_index=True) if new_nr_data else pd.DataFrame()
            st.session_state.bbu_data = pd.concat(new_bbu_data, ignore_index=True) if new_bbu_data else pd.DataFrame()
            self.update_status(f"Loaded {len(files)} files")
            st.success("Data loading completed!")
        except Exception as e:
            logging.error(f"Error in load_data: {str(e)}")
            self.update_status(f"Error loading files: {str(e)}")
            st.error(f"Error loading files: {str(e)}")

    def get_column_value(self, record, possible_names, tech):
        for name in possible_names:
            if name in record and pd.notna(record[name]):
                return str(record[name])
        if tech == "5GNR" and possible_names[0] == "CONFIGURATION":
            usid = self.get_column_value(record, self.mappings["5GNR"]["USID"], "5GNR")
            nrcell = self.get_column_value(record, self.mappings["5GNR"]["cell"], "5GNR")
            if usid and nrcell:
                key = f"{usid}_{nrcell}"
                bbu_row = st.session_state.bbu_data[
                    (st.session_state.bbu_data[self.mappings["5GNR_BBU"]["USID"]].apply(lambda x: str(x)) == usid) &
                    (st.session_state.bbu_data[self.mappings["5GNR_BBU"]["NRCELL_NAME"]].apply(lambda x: str(x)) == nrcell)
                ]
                if not bbu_row.empty:
                    for name in self.mappings["5GNR_BBU"]["CONFIGURATION"]:
                        if name in bbu_row.iloc[0] and pd.notna(bbu_row.iloc[0][name]):
                            return str(bbu_row.iloc[0][name])
        return ""

    def perform_search(self, search_type, search_value):
        if not search_value:
            st.error("Please enter a search value")
            return
        new_matched_records = []
        def search_in_data(data, tech):
            mapping = self.mappings[tech][search_type]
            for _, record in data.iterrows():
                for col_name in mapping:
                    if col_name in record and pd.notna(record[col_name]) and search_value.lower() in str(record[col_name]).lower():
                        new_matched_records.append((tech, record.to_dict()))
                        break
        if not st.session_state.lte_data.empty:
            search_in_data(st.session_state.lte_data, "LTE")
        if not st.session_state.nr_data.empty:
            search_in_data(st.session_state.nr_data, "5GNR")
        st.session_state.matched_records = new_matched_records
        if st.session_state.auto_generate:
            self.generate_vdt_data()
        self.update_status(f"Found {len(new_matched_records)} matching records")
        st.success(f"Found {len(new_matched_records)} matching records")

    def generate_vdt_data(self):
        lte_sites = sorted(set(self.get_column_value(record, self.mappings["LTE"]["Site"], "LTE")
                             for tech, record in st.session_state.matched_records if tech == "LTE" and self.get_column_value(record, self.mappings["LTE"]["Site"], "LTE")))
        nr_sites = sorted(set(self.get_column_value(record, self.mappings["5GNR"]["Site"], "5GNR")
                            for tech, record in st.session_state.matched_records if tech == "5GNR" and self.get_column_value(record, self.mappings["5GNR"]["Site"], "5GNR")))
        max_len = max(len(lte_sites), len(nr_sites))
        vdt_rows = [{"LTE Site": lte_sites[i] if i < len(lte_sites) else "", "NR Site": nr_sites[i] if i < len(nr_sites) else ""} for i in range(max_len)]
        st.session_state.vdt_data = pd.DataFrame(vdt_rows)
        self.update_status(f"Generated VDT data with {len(vdt_rows)} rows")

    def export_to_excel(self, df, filename):
        try:
            output = io.BytesIO()
            df.to_excel(output, index=False)
            output.seek(0)
            st.download_button(label=f"Download {filename}", data=output, file_name=filename,
                             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.update_status(f"Exported {len(df)} rows to {filename}")
        except Exception as e:
            logging.error(f"Error in export_to_excel: {str(e)}")
            st.error(f"Export failed: {str(e)}")

    def export_lte_to_excel(self):
        lte_data = [{"Source": "LTE", "Site": self.get_column_value(record, self.mappings["LTE"]["Site"], "LTE"),
                    "cell": self.get_column_value(record, self.mappings["LTE"]["cell"], "LTE"),
                    "CELLRANGE": self.get_column_value(record, self.mappings["LTE"]["CELLRANGE"], "LTE"),
                    "CRSGAIN": self.get_column_value(record, self.mappings["LTE"]["CRSGAIN"], "LTE"),
                    "QRXLEVMIN": self.get_column_value(record, self.mappings["LTE"]["QRXLEVMIN"], "LTE"),
                    "EARFCNDL": self.get_column_value(record, self.mappings["LTE"]["EARFCNDL"], "LTE")}
                   for tech, record in st.session_state.matched_records if tech == "LTE"]
        if not lte_data:
            st.error("No LTE data to export")
            return
        df = pd.DataFrame(lte_data, columns=st.session_state.lte_columns)
        self.export_to_excel(df, "LTE_Parameters.xlsx")

    def export_5g_to_excel(self):
        nr_data = [{"Source": "5GNR", "USID": self.get_column_value(record, self.mappings["5GNR"]["USID"], "5GNR"),
                   "SITE": self.get_column_value(record, self.mappings["5GNR"]["Site"], "5GNR"),
                   "NRCELL_NAME": self.get_column_value(record, self.mappings["5GNR"]["cell"], "5GNR"),
                   "Digital Tilt": self.get_column_value(record, self.mappings["5GNR"]["Digital Tilt"], "5GNR"),
                   "Power": self.get_column_value(record, self.mappings["5GNR"]["Power"], "5GNR"),
                   "PCI": self.get_column_value(record, self.mappings["5GNR"]["PCI"], "5GNR"),
                   "ADMINISTRATIVESTATE": self.get_column_value(record, self.mappings["5GNR"]["ADMINISTRATIVESTATE"], "5GNR"),
                   "CELLBARRED": self.get_column_value(record, self.mappings["5GNR"]["CELLBARRED"], "5GNR"),
                   "CELLRESERVEDFOROPERATOR": self.get_column_value(record, self.mappings["5GNR"]["CELLRESERVEDFOROPERATOR"], "5GNR"),
                   "OPERATIONALSTATE": self.get_column_value(record, self.mappings["5GNR"]["OPERATIONALSTATE"], "5GNR"),
                   "CELLRANGE": self.get_column_value(record, self.mappings["5GNR"]["CELLRANGE"], "5GNR"),
                   "SSBFREQUENCY": self.get_column_value(record, self.mappings["5GNR"]["SSBFREQUENCY"], "5GNR"),
                   "CONFIGURATION": self.get_column_value(record, self.mappings["5GNR"]["CONFIGURATION"], "5GNR")}
                  for tech, record in st.session_state.matched_records if tech == "5GNR"]
        if not nr_data:
            st.error("No 5G data to export")
            return
        df = pd.DataFrame(nr_data, columns=st.session_state.nr_columns)
        self.export_to_excel(df, "5G_Parameters.xlsx")

    def export_lte_selected_to_excel(self, df):
        if df.empty:
            st.error("No rows selected for export")
            return
        self.export_to_excel(df, "LTE_Selected.xlsx")

    def export_5g_selected_to_excel(self, df):
        if df.empty:
            st.error("No rows selected for export")
            return
        self.export_to_excel(df, "5G_Selected.xlsx")

    def generate_lte_cr(self, lte_cr_type):
        lte_data = [{"Source": "LTE", "Site": self.get_column_value(record, self.mappings["LTE"]["Site"], "LTE"),
                    "cell": self.get_column_value(record, self.mappings["LTE"]["cell"], "LTE"),
                    "CELLRANGE": self.get_column_value(record, self.mappings["LTE"]["CELLRANGE"], "LTE"),
                    "CRSGAIN": self.get_column_value(record, self.mappings["LTE"]["CRSGAIN"], "LTE"),
                    "QRXLEVMIN": self.get_column_value(record, self.mappings["LTE"]["QRXLEVMIN"], "LTE"),
                    "EARFCNDL": self.get_column_value(record, self.mappings["LTE"]["EARFCNDL"], "LTE"),
                    "Electrical Tilt": self.get_column_value(record, self.mappings["LTE"]["Electrical Tilt"], "LTE")}
                   for tech, record in st.session_state.matched_records if tech == "LTE"]
        if not lte_data:
            st.error("No LTE rows selected")
            return
        cr_data = []
        for row in lte_data:
            site = row["Site"]
            cell = row["cell"]
            value = ""
            mo_class = f"EUtranCellFDD={cell}"
            if lte_cr_type == "cellRange":
                value = row["CELLRANGE"]
            elif lte_cr_type == "crsGain":
                value = row["CRSGAIN"]
            elif lte_cr_type == "Electrical Tilt":
                value = row["Electrical Tilt"]
            cr_data.append({"Site": site, "MO Class": mo_class, "Parameter": lte_cr_type, "Value": value, "CurrentValue": ""})
        if cr_data:
            df = pd.DataFrame(cr_data)
            self.export_to_excel(df, "LTE_CR.xlsx")
            self.update_status(f"Generated {len(cr_data)} LTE CRs for {lte_cr_type}")

    def generate_5g_cr(self, nr_cr_type):
        nr_data = [{"Source": "5GNR", "USID": self.get_column_value(record, self.mappings["5GNR"]["USID"], "5GNR"),
                   "SITE": self.get_column_value(record, self.mappings["5GNR"]["Site"], "5GNR"),
                   "NRCELL_NAME": self.get_column_value(record, self.mappings["5GNR"]["cell"], "5GNR"),
                   "Digital Tilt": self.get_column_value(record, self.mappings["5GNR"]["Digital Tilt"], "5GNR"),
                   "Power": self.get_column_value(record, self.mappings["5GNR"]["Power"], "5GNR"),
                   "PCI": self.get_column_value(record, self.mappings["5GNR"]["PCI"], "5GNR"),
                   "CELLRANGE": self.get_column_value(record, self.mappings["5GNR"]["CELLRANGE"], "5GNR")}
                  for tech, record in st.session_state.matched_records if tech == "5GNR"]
        if not nr_data:
            st.error("No 5G rows selected")
            return
        cr_data = []
        for row in nr_data:
            site = row["SITE"]
            cell = row["NRCELL_NAME"]
            value = ""
            mo_class = f"NRCellDU={cell}" if nr_cr_type == "cellRange" else f"NRSectorCarrier={cell},CommonBeamforming=1"
            if nr_cr_type == "digitalTilt":
                value = row["Digital Tilt"]
            elif nr_cr_type == "cellRange":
                value = row["CELLRANGE"]
            cr_data.append({"Site": site, "MO Class": mo_class, "Parameter": nr_cr_type, "Value": value, "CurrentValue": ""})
        if cr_data:
            df = pd.DataFrame(cr_data)
            self.export_to_excel(df, "5G_CR.xlsx")
            self.update_status(f"Generated {len(cr_data)} 5G CRs for {nr_cr_type}")

    def generate_vdt_report(self, project_name):
        lte_sites = sorted(set(self.get_column_value(record, self.mappings["LTE"]["Site"], "LTE")
                             for tech, record in st.session_state.matched_records if tech == "LTE" and self.get_column_value(record, self.mappings["LTE"]["Site"], "LTE")))
        nr_sites = sorted(set(self.get_column_value(record, self.mappings["5GNR"]["Site"], "5GNR")
                            for tech, record in st.session_state.matched_records if tech == "5GNR" and self.get_column_value(record, self.mappings["5GNR"]["Site"], "5GNR")))
        if not lte_sites and not nr_sites:
            st.error("No sites found in VDT tab. Please generate VDT data first.")
            return
        wb = openpyxl.Workbook()
        # LTE Sheet
        lte_data = [
            ["Type", "Value", "Remarks", "Buffer", ""],
            ["projectName", project_name, "", "0.01", ""],
            ["startTime", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Use exact format. Add \" ' \" before", "Default keep it 0.01. Can increase it to accommodate TA Plot", ""],
            ["endTime", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Use exact format. Add \" ' \" before", "", ""],
            ["Site Name List", "ZOOM (Default = 0 for automatic zoom, enter integer value for manual zoom\nEg. value: 5 will fetch plots up to 5 km diagonally from site in both directions.)",
             "Neighbour list (comma separated), if blank first tier neighbours will be considered", "Carrier List(comma separated)", ""]
        ]
        for site in lte_sites:
            earfcns = sorted(set(self.get_column_value(record, self.mappings["LTE"]["EARFCNDL"], "LTE")
                               for tech, record in st.session_state.matched_records
                               if tech == "LTE" and self.get_column_value(record, self.mappings["LTE"]["Site"], "LTE") == site
                               and self.get_column_value(record, self.mappings["LTE"]["EARFCNDL"], "LTE")))
            lte_data.append([site, "0", "", ",".join(earfcns), ""])
        ws_lte = wb.create_sheet("LTE")
        for row in lte_data:
            ws_lte.append(row)
        # NR Sheet
        nr_data = [
            ["Type", "Value", "Remarks", "Buffer", "", ""],
            ["projectName", project_name, "", "0.01", "", ""],
            ["startTime", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Use exact format. Add \" ' \" before", "Default keep it 0.01. Can increase it to accommodate TA Plot", "", ""],
            ["endTime", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Use exact format. Add \" ' \" before", "", "", ""],
            ["Site Name List", "ZOOM (Default = 0 for automatic zoom, enter integer value for manual zoom\nEg. value: 5 will fetch plots up to 5 km diagonally from site in both directions.)",
             "Neighbour list (comma separated), if blank first tier neighbours will be considered", "Carrier List(comma separated)", "", ""]
        ]
        for site in nr_sites:
            arfcns = sorted(set(self.get_column_value(record, self.mappings["5GNR"]["SSBFREQUENCY"], "5GNR")
                              for tech, record in st.session_state.matched_records
                              if tech == "5GNR" and self.get_column_value(record, self.mappings["5GNR"]["Site"], "5GNR") == site
                              and self.get_column_value(record, self.mappings["5GNR"]["SSBFREQUENCY"], "5GNR")))
            nr_data.append([site, "0", "", ",".join(arfcns), "", ""])
        ws_nr = wb.create_sheet("NR")
        for row in nr_data:
            ws_nr.append(row)
        wb.remove(wb["Sheet"])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        st.download_button(label="Download VDT Report", data=output, file_name="VDT_Report.xlsx",
                         mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.update_status(f"VDT report generated with {len(lte_sites)} LTE and {len(nr_sites)} NR sites")

    def add_point(self, name, lat, lon):
        try:
            if not name:
                name = f"Point {len(st.session_state.points) + 1}"
            if not lat or not lon:
                st.error("Please enter both latitude and longitude")
                return
            st.session_state.points.append((name, float(lat), float(lon)))
            self.update_status(f"Added point {name}")
        except Exception as e:
            logging.error(f"Error in add_point: {str(e)}")
            st.error(f"Invalid latitude or longitude: {str(e)}")

    def add_lat_long(self, lat_long, name):
        try:
            lat_long = lat_long.split(",")
            if len(lat_long) != 2:
                st.error("Invalid format. Use 'latitude,longitude'")
                return
            lat, lon = map(float, lat_long)
            self.add_point(name, lat, lon)
        except Exception as e:
            logging.error(f"Error in add_lat_long: {str(e)}")
            st.error(f"Invalid format: {str(e)}")

    def import_points_from_excel(self, file):
        try:
            df = pd.read_excel(file)
            new_points = []
            for _, row in df.iterrows():
                name = row.get("name") or row.get("Name") or row.get("point") or row.get("Point") or f"Point {len(st.session_state.points) + len(new_points) + 1}"
                lat = row.get("lat") or row.get("Latitude") or row.get("LAT")
                lon = row.get("lon") or row.get("Longitude") or row.get("LON") or row.get("long") or row.get("Long")
                if lat is not None and lon is not None:
                    new_points.append((name, float(lat), float(lon)))
            st.session_state.points.extend(new_points)
            self.update_status(f"Imported {len(new_points)} points")
        except Exception as e:
            logging.error(f"Error in import_points_from_excel: {str(e)}")
            st.error(f"Invalid Excel file: {str(e)}")

    def set_master_point(self, point):
        st.session_state.master_point = point
        self.update_status(f"Set master point: {point[0]}")

    def remove_points(self, selected):
        if selected:
            st.session_state.points = [p for p in st.session_state.points if p[0] not in selected]
            self.update_status(f"Removed {len(selected)} points")

    def clear_points(self):
        st.session_state.points = []
        st.session_state.master_point = None
        self.update_status("Cleared all points")

    def calculate_path_distances(self):
        if len(st.session_state.points) < 2:
            st.error("Need at least 2 points to calculate distances")
            return
        results = []
        for i in range(len(st.session_state.points)):
            for j in range(i + 1, len(st.session_state.points)):
                p1, p2 = st.session_state.points[i], st.session_state.points[j]
                dist = self.haversine_distance(p1[1], p1[2], p2[1], p2[2])
                results.append(f"{p1[0]} to {p2[0]}: {dist:.2f} km")
        st.session_state.distance_results = "\n".join(results)
        self.update_status(f"Calculated distances for {len(results)} pairs")

    def calculate_from_master(self):
        if not st.session_state.master_point or not st.session_state.points:
            st.error("No master point or points set")
            return
        results = []
        master = st.session_state.master_point
        for point in st.session_state.points:
            if point != master:
                dist = self.haversine_distance(master[1], master[2], point[1], point[2])
                results.append(f"{master[0]} to {point[0]}: {dist:.2f} km")
        st.session_state.distance_results = "\n".join(results)
        self.update_status(f"Calculated distances from master point {master[0]}")

    def haversine_distance(self, lat1, lon1, lat2, lon2):
        R = 6371  # Earth radius in kilometers
        lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
        dlat = lat2 - lat1
        dlon = lon2 - lon1
        a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
        c = 2 * atan2(sqrt(a), sqrt(1-a))
        return R * c

    def show_map(self, map_type, map_zoom):
        if not st.session_state.points:
            st.error("No points to display on map")
            return
        points_str = "&markers=".join([f"{lat},{lon}" for _, lat, lon in st.session_state.points])
        if st.session_state.master_point:
            master_lat, master_lon = st.session_state.master_point[1], st.session_state.master_point[2]
            points_str += f"&markers=color:blue%7Clabel:M%7C{master_lat},{master_lon}"
        center_lat = sum(p[1] for p in st.session_state.points) / len(st.session_state.points)
        center_lon = sum(p[2] for p in st.session_state.points) / len(st.session_state.points)
        base_url = f"https://www.google.com/maps/embed/v1/view?key={self.api_key}&center={center_lat},{center_lon}&zoom={map_zoom}&maptype={map_type}&markers={points_str}"
        st.components.v1.iframe(base_url, height=400, scrolling=False)

    def add_lte_column(self):
        try:
            data_sources = ["LTE"]
            selected_source = "LTE"
            column_name = st.selectbox("Select Column:", list(self.mappings[selected_source].keys()), key="lte_add_column")
            if column_name and column_name not in st.session_state.lte_columns:
                st.session_state.lte_columns.append(column_name)
                self.update_status(f"Added column {column_name} to LTE Parameters")
        except Exception as e:
            logging.error(f"Error in add_lte_column: {str(e)}")
            st.error(f"Failed to add column: {str(e)}")

    def add_5g_column(self):
        try:
            data_sources = ["5GNR", "5GNR_BBU"]
            selected_source = st.selectbox("Select Data Source:", data_sources, key="5g_source")
            column_name = st.selectbox("Select Column:", list(self.mappings[selected_source].keys()), key="5g_add_column")
            if column_name and column_name not in st.session_state.nr_columns:
                st.session_state.nr_columns.append(column_name)
                self.update_status(f"Added column {column_name} to 5G Parameters")
        except Exception as e:
            logging.error(f"Error in add_5g_column: {str(e)}")
            st.error(f"Failed to add column: {str(e)}")

    def use_for_distance(self, df):
        try:
            if 'LATITUDE' in df.columns and 'LONGITUDE' in df.columns:
                for _, row in df.iterrows():
                    lat = float(row['LATITUDE']) if pd.notna(row['LATITUDE']) else None
                    lon = float(row['LONGITUDE']) if pd.notna(row['LONGITUDE']) else None
                    if lat is not None and lon is not None:
                        name = row.get('Site', f"Point {len(st.session_state.points) + 1}")
                        self.add_point(name, lat, lon)
                self.update_status(f"Added {len(df)} points from search results")
            else:
                st.error("Selected data must contain LATITUDE and LONGITUDE columns")
        except Exception as e:
            logging.error(f"Error in use_for_distance: {str(e)}")
            st.error(f"Failed to use for distance: {str(e)}")

    def copy_with_headers(self, df):
        try:
            if df.empty:
                st.error("No data to copy")
                return
            header_text = "\t".join(df.columns)
            data_rows = "\n".join(["\t".join(map(str, row)) for row in df.values])
            clipboard_text = header_text + "\n" + data_rows
            st.session_state.clipboard = clipboard_text
            st.success("Data copied to clipboard")
            self.update_status("Copied data with headers to clipboard")
        except Exception as e:
            logging.error(f"Error in copy_with_headers: {str(e)}")
            st.error(f"Failed to copy: {str(e)}")

    def clear_results(self):
        st.session_state.matched_records = []
        st.session_state.lte_tree_record_map = {}
        st.session_state.nr_tree_record_map = {}
        self.update_status("Cleared all results")

def main():
    try:
        app = NetworkSearchApp()
    except Exception as e:
        logging.critical(f"Critical Error: {str(e)}\n{traceback.format_exc()}")
        st.error(f"Application crashed: {str(e)}")

if __name__ == "__main__":
    main()