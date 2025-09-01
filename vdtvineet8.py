import pandas as pd
import os
import sys
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext, simpledialog
import traceback
from datetime import datetime
import openpyxl
import math
from math import radians, sin, cos, sqrt, atan2
import webbrowser
from PIL import Image, ImageTk
import io
import requests
import urllib.parse
import tempfile
import logging
import difflib

# Configure logging
logging.basicConfig(filename='network_search.log', level=logging.ERROR,
                    format='%(asctime)s - %(levelname)s - %(message)s')

class NetworkSearchApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Network Data Search Tool")
        self.root.geometry("1400x800")
        self.root.state('zoomed')  # Start maximized
        
        # Initialize data structures
        self.lte_data = pd.DataFrame()
        self.nr_data = pd.DataFrame()
        self.bbu_data = pd.DataFrame()
        self.usid_index = {}
        self.bbu_index = {}
        self.file_paths = {}
        self.matched_records = []
        self.points = []
        self.master_point = None
        
        # VDT data
        self.vdt_data = pd.DataFrame(columns=["LTE Site", "NR Site"])
        self.lte_tree_record_map = {}  # For VDT data mapping
        self.nr_tree_record_map = {}   # For VDT data mapping
        self.auto_generate_var = tk.BooleanVar(value=True)  # Auto-generate VDT
        
        # Market mapping
        self.market_mapping = {
            "ATT_ARK1": "ATT_ARK_253",
            "ATT_NoCAL1": "ATT_NoCAL_253",
            "ATT_SoCAL1": "ATT_SoCAL_253",
            "ATT_STX": "ATT_STX_253"
        }
        self.project_name_var = tk.StringVar(value="ATT_STX_253")
        
        # Enhanced mappings with better 5G support and additional mappings
        self.mappings = {
            "LTE": {
                "USID": ["REMOTE_USID", "CSS_USID", "USID"],
                "ENBID": ["ENBID"],
                "cell ID": ["CELLID"],
                "Site": ["MECONTEXT_ID", "SITE", "OSS_ENodeB"],
                "Azumuth": ["ATOLL_AZIMUTH", "AZIMUTH", "Atoll_AZIMUT", "Atoll_Az"],
                "Digital Tilt": ["DIGITALTILT", "DIGITAL_TILT", "TILT"],
                "cell": ["EUTRAN_CELL_FDD_ID", "CELL", "OSS_EUTRAN_CELL_FDD_ID", "CELL_NAME"],
                "height(Meter)": ["Atoll_HEIGHT_m", "HEIGHT", "ANTENNA_HEIGHT", "HEIGHT_M"],
                "PCI": ["PHYSICALLAYERCELLID", "PCI", "OSS_PCI", "Atoll_PCI"],
                "Power": ["CONFIGUREDMAXTXPOWER", "TX_POWER", "OSS_CONFIGUREDMAXTXPOWER"],
                "LATITUDE": ["LATITUDE", "LAT"],
                "LONGITUDE": ["LONGITUDE", "LON", "LONG"],
                "ADMINISTRATIVESTATE": ["ADMINISTRATIVE_STATE", "ADMIN_STATE"],
                "OPERATIONALSTATE": ["OPERATIONALSTATE", "OP_STATE"],
                "CELLRANGE": ["CELLRANGE"],
                "CRSGAIN": ["CRSGAIN"],
                "QRXLEVMIN": ["QRXLEVMIN"],
                "EARFCNDL": ["EARFCNDL"],
                "Electrical Tilt": ["ELECTRICAL_TILT", "Atoll_ET", "E_TILT"],
                "ED_Market": ["ED_MARKET", "ED_Market", "EDMARKET"]
            },
            "5GNR": {
                "USID": ["CSS_USID", "REMOTE_USID", "USID", "BBU_USID"],
                "NIC": ["NCI"],
                "gnb ID": ["GNBID", "GNB_ID"],
                "Site": [ "CTS_COMMON_ID", "GNB_NAME", "GNODEB"],
                "Azumuth": ["Atoll_AZIMUT", "AZIMUTH", "ATOLL_AZIMUTH"],
                "Digital Tilt": ["DIGITALTILT", "DIGITAL_TILT", "USEDDIGITALTILT"],
                "cell": ["NRCELLDUID", "CELL", "NRCELL_NAME"],
                "height(Meter)": ["Atoll_HEIGHT_m", "HEIGHT", "ANTENNA_HEIGHT", "ATOLL_ANTENNA_HEIGHT"],
                "PCI": ["NRPCI", "PCI"],
                "Power": ["CONFIGUREDMAXTXPOWER", "TX_POWER", "POWER"],
                "LATITUDE": ["LAT", "LATITUDE"],
                "LONGITUDE": ["LONG", "LONGITUDE"],
                "ADMINISTRATIVESTATE": ["ADMINISTRATIVESTATE", "ADMIN_STATE"],
                "OPERATIONALSTATE": ["OPERATIONALSTATE", "OP_STATE"],
                "CELLBARRED": ["CELLBARRED"],
                "CELLRESERVEDFOROPERATOR": ["CELLRESERVEDFOROPERATOR"],
                "CELLRANGE": ["CELLRANGE"],
                "SSBFREQUENCY": ["SSBFREQUENCY"],
                "ARFCNDL": ["ARFCNDL", "ARFCN_DL", "DL_ARFCN", "NR_ARFCNDL"],
                "CONFIGURATION": ["CONFIGURATION"],
                "Electrical Tilt": ["ELECTRICAL_TILT", "Atoll_ET", "E_TILT"],
                "ED_Market": ["ED_MARKET", "ED_Market", "EDMARKET"],
                "BBU_TECH": ["BBU_TECH"],
                "GNB_SA_STATE": ["GNB_SA_STATE"],
                "CELL_SA_STATE": ["CELL_SA_STATE"],
                "CELL_TYPE": ["CELL_TYPE"],
                "ON_AIR": ["ON_AIR"],
                "DSS_LTECELL": ["DSS_LTECELL"],
                "NRTAC": ["NRTAC"]
            },
            "5GNR_BBU": {
                "USID": ["USID", "REMOTE_USID", "CSS_USID"],
                "NRCELL_NAME": ["NRCELLDUID", "CELL", "NRCELL_NAME"],
                "CONFIGURATION": ["CONFIGURATION"],
                "ED_Market": ["ED_MARKET", "ED_Market", "EDMARKET"]
            }
        }
        
        # Google Maps API Key
        self.api_key = os.environ.get("GOOGLE_MAPS_API_KEY", "")
        
        # Create UI
        self.create_widgets()
        
        # Status bar
        self.status_var = tk.StringVar()
        status_bar = ttk.Label(root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        self.update_status("Ready. Please load data files.")

        # Configure Treeview style for smaller font
        self.style = ttk.Style()
        self.style.configure("Treeview", font=('Arial', 8), rowheight=20)
        self.style.configure("Treeview.Heading", font=('Arial', 9, 'bold'))
        
        # Context menu for 5G rows
        self.nr_context_menu = tk.Menu(self.root, tearoff=0)
        self.nr_context_menu.add_command(label="Copy Cell", command=self.copy_nr_cell)
        self.nr_context_menu.add_command(label="Copy with Headers", command=self.copy_5g_with_headers)
        self.nr_context_menu.add_command(label="Export Selected to Excel", command=self.export_5g_selected_to_excel)
        self.nr_context_menu.add_command(label="Clear Results", command=self.clear_results)
        
        # Context menu for LTE rows
        self.lte_context_menu = tk.Menu(self.root, tearoff=0)
        self.lte_context_menu.add_command(label="Copy Cell", command=self.copy_lte_cell)
        self.lte_context_menu.add_command(label="Copy with Headers", command=self.copy_lte_with_headers)
        self.lte_context_menu.add_command(label="Export Selected to Excel", command=self.export_lte_selected_to_excel)
        self.lte_context_menu.add_command(label="Clear Results", command=self.clear_results)
    
    def update_status(self, message):
        """Update status bar with timestamp"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.status_var.set(f"[{timestamp}] {message}")
    
    def create_widgets(self):
        # Main frame with padding
        main_frame = ttk.Frame(self.root, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Header
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=5)
        
        title_label = ttk.Label(header_frame, text="Network Data Search Tool", 
                               font=("Arial", 16, "bold"))
        title_label.pack(side=tk.LEFT)
        
        # File loading section - made more compact
        file_frame = ttk.LabelFrame(main_frame, text="Data Loading", padding=5)  # Reduced padding
        file_frame.pack(fill=tk.X, pady=3)  # Reduced padding
        
        # Multi-file upload
        upload_frame = ttk.Frame(file_frame)
        upload_frame.pack(fill=tk.X, pady=3)  # Reduced padding
        
        ttk.Label(upload_frame, text="Upload Multiple Files:").pack(side=tk.LEFT, padx=5)
        self.file_listbox = tk.Listbox(upload_frame, width=80, height=2, selectmode=tk.MULTIPLE)  # Reduced height
        self.file_listbox.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        upload_btn = ttk.Button(upload_frame, text="Browse Files", command=self.browse_files)
        upload_btn.pack(side=tk.LEFT, padx=5)
        
        # Load button
        load_btn = ttk.Button(file_frame, text="Load Selected Data", command=self.load_data)
        load_btn.pack(pady=5)  # Reduced padding
        
        # Search section
        search_frame = ttk.LabelFrame(main_frame, text="Search", padding=10)
        search_frame.pack(fill=tk.X, pady=5)
        
        search_control_frame = ttk.Frame(search_frame)
        search_control_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(search_control_frame, text="Search By:").pack(side=tk.LEFT, padx=5)
        self.search_type = ttk.Combobox(search_control_frame, 
                                      values=["USID", "NIC", "gnb ID", "ENBID", "cell ID", "Site"],
                                      width=10, state="readonly")
        self.search_type.current(0)
        self.search_type.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(search_control_frame, text="Value:").pack(side=tk.LEFT, padx=5)
        self.search_entry = ttk.Entry(search_control_frame, width=30)
        self.search_entry.pack(side=tk.LEFT, padx=5)
        
        search_btn = ttk.Button(search_control_frame, text="Search", command=self.perform_search)
        search_btn.pack(side=tk.LEFT, padx=10)
        
        # Create notebook for results tabs
        self.results_notebook = ttk.Notebook(main_frame)
        self.results_notebook.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Create main results tab
        main_tab = ttk.Frame(self.results_notebook)
        self.results_notebook.add(main_tab, text="Main Results")
        self.create_main_tab(main_tab)
        
        # Create LTE parameters tab
        lte_tab = ttk.Frame(self.results_notebook)
        self.results_notebook.add(lte_tab, text="LTE Parameters")
        self.create_lte_tab(lte_tab)
        
        # Create 5G Parameters tab
        nr_tab = ttk.Frame(self.results_notebook)
        self.results_notebook.add(nr_tab, text="5G Parameters")
        self.create_5g_tab(nr_tab)
        
        # Create VDT Sheet tab
        vdt_tab = ttk.Frame(self.results_notebook)
        self.results_notebook.add(vdt_tab, text="VDT Sheet")
        self.create_vdt_tab(vdt_tab)
        
        # Create Distance Calculator tab
        dist_tab = ttk.Frame(self.results_notebook)
        self.results_notebook.add(dist_tab, text="Distance Calculator")
        self.create_distance_tab(dist_tab)
    
    def create_main_tab(self, parent):
        """Create the main results tab"""
        results_frame = ttk.LabelFrame(parent, text="Main Results", padding=10)
        results_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create a Treeview with scrollbars
        columns = (
            "Source", "NIC", "gnb ID", "ENBID", "cell ID", "USID", "Site", 
            "Azumuth", "Digital Tilt", "cell", "height(Meter)", "PCI", "Power", 
            "LATITUDE", "LONGITUDE", "ADMINISTRATIVESTATE", "OPERATIONALSTATE"
        )
        
        self.tree = ttk.Treeview(results_frame, columns=columns, show="headings", selectmode="extended")
        vsb = ttk.Scrollbar(results_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(results_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Configure columns with headings
        col_widths = {
            "Source": 50, "NIC": 70, "gnb ID": 50, "ENBID": 50, 
            "cell ID": 50, "USID": 70, "Site": 70, "Azumuth": 50,
            "Digital Tilt": 70, "cell": 90, "height(Meter)": 70, "PCI": 40,
            "Power": 50, "LATITUDE": 70, "LONGITUDE": 70,
            "ADMINISTRATIVESTATE": 90, "OPERATIONALSTATE": 90
        }
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=col_widths.get(col, 80), anchor=tk.CENTER)
        
        # Layout treeview and scrollbars
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        # Configure grid weights
        results_frame.grid_rowconfigure(0, weight=1)
        results_frame.grid_columnconfigure(0, weight=1)
        
        # Create tag for highlighted cells
        self.tree.tag_configure('highlight', background='#f0f0f0')
        
        # Add export button
        export_frame = ttk.Frame(results_frame)
        export_frame.grid(row=2, column=0, sticky="e", pady=5)
        export_btn = ttk.Button(export_frame, text="Export to Excel", command=self.export_to_excel)
        export_btn.pack(side=tk.RIGHT, padx=5)
        
        # Add context menu
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="Copy with Headers", command=self.copy_with_headers)
        self.context_menu.add_command(label="Export Selected to Excel", command=self.export_selected_to_excel)
        self.context_menu.add_command(label="Use for Distance Calculation", command=self.use_for_distance)
        self.context_menu.add_command(label="Clear Results", command=self.clear_results)
        self.tree.bind("<Button-3>", self.show_context_menu)
    
    def create_lte_tab(self, parent):
        """Create the LTE parameters tab with added Site column"""
        results_frame = ttk.LabelFrame(parent, text="LTE Parameters", padding=10)
        results_frame.pack(fill=tk.BOTH, expand=True)
        
        # Control frame for buttons and dropdowns
        control_frame = ttk.Frame(results_frame)
        control_frame.pack(fill=tk.X, pady=5)
        
        # Add Column button
        add_col_btn = ttk.Button(control_frame, text="Add Column", command=self.add_lte_column)
        add_col_btn.pack(side=tk.LEFT, padx=5)
        
        # CR Generation section
        ttk.Label(control_frame, text="Generate CR for:").pack(side=tk.LEFT, padx=5)
        self.lte_cr_type = ttk.Combobox(control_frame, 
                                       values=["cellRange", "crsGain", "Electrical Tilt"],
                                       width=15, state="readonly")
        self.lte_cr_type.current(0)
        self.lte_cr_type.pack(side=tk.LEFT, padx=5)
        
        # Generate CR button
        cr_btn = ttk.Button(control_frame, text="Generate CR", command=self.generate_lte_cr)
        cr_btn.pack(side=tk.LEFT, padx=5)
        
        # Export LTE Data to Excel button
        export_btn = ttk.Button(control_frame, text="Export LTE Data to Excel", 
                               command=self.export_lte_to_excel)
        export_btn.pack(side=tk.LEFT, padx=5)
        
        # Create a container frame for the treeview and scrollbars
        tree_container = ttk.Frame(results_frame)
        tree_container.pack(fill=tk.BOTH, expand=True)
        
        # Create Treeview with scrollbars
        columns = (
            "Source", "Site", "cell", "CELLRANGE", "CRSGAIN", "QRXLEVMIN", "EARFCNDL"
        )
        
        self.lte_tree = ttk.Treeview(tree_container, columns=columns, show="headings", selectmode="extended")
        vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self.lte_tree.yview)
        hsb = ttk.Scrollbar(tree_container, orient="horizontal", command=self.lte_tree.xview)
        self.lte_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Configure columns with headings
        col_widths = {
            "Source": 50, "Site": 70, "cell": 90, "CELLRANGE": 50, 
            "CRSGAIN": 50, "QRXLEVMIN": 60, "EARFCNDL": 50
        }
        
        for col in columns:
            self.lte_tree.heading(col, text=col)
            self.lte_tree.column(col, width=col_widths.get(col, 80), anchor=tk.CENTER)
        
        # Layout treeview and scrollbars
        self.lte_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        # Configure grid weights
        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)
        
        # Bind right-click
        self.lte_tree.bind("<Button-3>", self.handle_lte_right_click)
    
    def handle_lte_right_click(self, event):
        """Handle right-click in LTE treeview"""
        try:
            region = self.lte_tree.identify("region", event.x, event.y)
            if region == "heading":
                col_id = self.lte_tree.identify_column(event.x)
                col_index = int(col_id.replace('#', '')) - 1
                columns = self.lte_tree['columns']
                if col_index < len(columns):
                    self.current_col_index = col_index
                    self.current_col_name = columns[col_index]
                    
                    # Create context menu for column
                    menu = tk.Menu(self.root, tearoff=0)
                    menu.add_command(label="Delete Column", command=self.delete_lte_column)
                    menu.add_command(label="Rename Column", command=self.rename_lte_column)
                    menu.tk_popup(event.x_root, event.y_root)
            else:
                # For rows/cells
                self.current_col = self.lte_tree.identify_column(event.x)
                self.current_row = self.lte_tree.identify_row(event.y)
                selection = self.lte_tree.selection()
                if selection:
                    self.lte_context_menu.entryconfig("Copy with Headers", state=tk.NORMAL)
                    self.lte_context_menu.entryconfig("Export Selected to Excel", state=tk.NORMAL)
                else:
                    self.lte_context_menu.entryconfig("Copy with Headers", state=tk.DISABLED)
                    self.lte_context_menu.entryconfig("Export Selected to Excel", state=tk.DISABLED)
                self.lte_context_menu.tk_popup(event.x_root, event.y_root)
        except Exception as e:
            logging.error(f"Error in handle_lte_right_click: {str(e)}")
    
    def copy_lte_cell(self):
        """Copy single cell value from LTE treeview"""
        try:
            if not hasattr(self, 'current_row') or not self.current_row:
                return
            item = self.current_row
            col_idx = int(self.current_col.replace('#', '')) - 1
            values = self.lte_tree.item(item, 'values')
            if col_idx < len(values):
                cell_value = str(values[col_idx])
                self.root.clipboard_clear()
                self.root.clipboard_append(cell_value)
                self.update_status("Copied cell value to clipboard")
        except Exception as e:
            logging.error(f"Error in copy_lte_cell: {str(e)}")
            messagebox.showerror("Error", f"Failed to copy cell: {str(e)}")
    
    def copy_lte_with_headers(self):
        """Copy selected rows with headers from LTE treeview"""
        try:
            selected = self.lte_tree.selection()
            if not selected:
                return
            headers = self.lte_tree['columns']
            header_text = "\t".join(headers)
            data_rows = []
            for item in selected:
                values = self.lte_tree.item(item, 'values')
                data_rows.append("\t".join(map(str, values)))
            clipboard_text = header_text + "\n" + "\n".join(data_rows)
            self.root.clipboard_clear()
            self.root.clipboard_append(clipboard_text)
            self.update_status("Copied selected rows with headers to clipboard")
        except Exception as e:
            logging.error(f"Error in copy_lte_with_headers: {str(e)}")
            messagebox.showerror("Error", f"Failed to copy: {str(e)}")
    
    def export_lte_selected_to_excel(self):
        """Export selected rows from LTE treeview to Excel"""
        try:
            selected = self.lte_tree.selection()
            if not selected:
                messagebox.showinfo("Info", "No rows selected for export")
                return
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if not file_path:
                return
            headers = self.lte_tree['columns']
            data = []
            for item in selected:
                data.append(self.lte_tree.item(item, 'values'))
            df = pd.DataFrame(data, columns=headers)
            df.to_excel(file_path, index=False)
            self.update_status(f"Exported {len(data)} rows to {os.path.basename(file_path)}")
            messagebox.showinfo("Success", "Data exported successfully!")
        except Exception as e:
            logging.error(f"Error in export_lte_selected_to_excel: {str(e)}")
            messagebox.showerror("Error", f"Export failed: {str(e)}")
    
    def add_lte_column(self):
        """Add a new column to the LTE treeview"""
        try:
            # Get available data sources
            data_sources = ["LTE", "5GNR", "5GNR_BBU"]
            
            # Create dialog for column selection
            dialog = tk.Toplevel(self.root)
            dialog.title("Add Column to LTE")
            dialog.geometry("400x300")
            dialog.transient(self.root)
            dialog.grab_set()
            
            # Source selection
            ttk.Label(dialog, text="Select Data Source:").pack(pady=5)
            source_var = tk.StringVar(value=data_sources[0])
            source_combo = ttk.Combobox(dialog, textvariable=source_var, values=data_sources, state="readonly")
            source_combo.pack(side=tk.LEFT, padx=5)
            
            # Column selection
            ttk.Label(dialog, text="Select Column:").pack(pady=5)
            column_listbox = tk.Listbox(dialog)
            column_listbox.pack(fill=tk.BOTH, expand=True, pady=5)
            
            # Populate initial columns
            self.populate_column_list(column_listbox, data_sources[0])
            
            # Update column list when source changes
            def on_source_change(event):
                self.populate_column_list(column_listbox, source_var.get())
            
            source_combo.bind("<<ComboboxSelected>>", on_source_change)
            
            # Add button
            def add_selected_column():
                selected = column_listbox.curselection()
                if selected:
                    column_name = column_listbox.get(selected[0])
                    current_columns = list(self.lte_tree['columns'])
                    
                    # Skip if column already exists
                    if column_name in current_columns:
                        messagebox.showinfo("Info", f"Column '{column_name}' already exists")
                        return
                    
                    # Add new column
                    current_columns.append(column_name)
                    self.rebuild_lte_tree(current_columns)
                    dialog.destroy()
            
            ttk.Button(dialog, text="Add", command=add_selected_column).pack(pady=10)
            
        except Exception as e:
            logging.error(f"Error in add_lte_column: {str(e)}")
            messagebox.showerror("Error", f"Failed to add column: {str(e)}")
    
    def populate_column_list(self, listbox, data_source):
        """Populate listbox with columns from specified data source"""
        listbox.delete(0, tk.END)
        columns = []
        if data_source == "LTE" and not self.lte_data.empty:
            columns = sorted(self.lte_data.columns)
        elif data_source == "5GNR" and not self.nr_data.empty:
            columns = sorted(self.nr_data.columns)
        elif data_source == "5GNR_BBU" and not self.bbu_data.empty:
            columns = sorted(self.bbu_data.columns)
        for col in columns:
            listbox.insert(tk.END, col)
    
    def rebuild_lte_tree(self, new_columns):
        """Rebuild the LTE treeview with new columns"""
        try:
            # Create a new tree with updated columns
            old_tree = self.lte_tree
            tree_container = old_tree.master  # This is the container frame
            results_frame = tree_container.master  # Parent of tree_container
            
            # Create new container
            new_container = ttk.Frame(results_frame)
            new_container.pack(fill=tk.BOTH, expand=True)
            
            # Create new tree
            self.lte_tree = ttk.Treeview(new_container, columns=new_columns, show="headings", selectmode="extended")
            vsb = ttk.Scrollbar(new_container, orient="vertical", command=self.lte_tree.yview)
            hsb = ttk.Scrollbar(new_container, orient="horizontal", command=self.lte_tree.xview)
            self.lte_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            
            # Configure columns
            col_widths = {
                "Source": 50, "Site": 70, "cell": 90, "CELLRANGE": 50, 
                "CRSGAIN": 50, "QRXLEVMIN": 60, "EARFCNDL": 50
            }
            
            for col in new_columns:
                self.lte_tree.heading(col, text=col)
                self.lte_tree.column(col, width=col_widths.get(col, 70), anchor=tk.CENTER)
            
            # Layout new tree and scrollbars
            self.lte_tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, sticky="ew")
            
            # Configure grid weights for container
            new_container.grid_rowconfigure(0, weight=1)
            new_container.grid_columnconfigure(0, weight=1)
            
            # Rebind right-click
            self.lte_tree.bind("<Button-3>", self.handle_lte_right_click)
            
            # Repopulate with existing records
            if hasattr(self, 'matched_records') and self.matched_records:
                for tech, record in self.matched_records:
                    if tech == "LTE":
                        self.add_to_lte_tree(tech, record)
            
            # Destroy old container and tree
            tree_container.destroy()
            
            self.update_status(f"LTE Parameters tab updated with {len(new_columns)} columns")
            
        except Exception as e:
            logging.error(f"Error in rebuild_lte_tree: {str(e)}")
            messagebox.showerror("Error", f"Failed to rebuild tree: {str(e)}")
    
    def delete_lte_column(self):
        """Delete selected column from LTE treeview"""
        try:
            if hasattr(self, 'current_col_name') and self.current_col_name:
                # Get current columns
                current_columns = list(self.lte_tree['columns'])
                
                # Skip required columns
                required_columns = {"Source", "Site", "cell"}
                if self.current_col_name in required_columns:
                    messagebox.showwarning("Cannot Delete", f"Cannot delete required column: {self.current_col_name}")
                    return
                
                # Remove the selected column
                if self.current_col_name in current_columns:
                    current_columns.remove(self.current_col_name)
                    
                    # Rebuild tree with updated columns
                    self.rebuild_lte_tree(current_columns)
        except Exception as e:
            logging.error(f"Error in delete_lte_column: {str(e)}")
            messagebox.showerror("Error", f"Failed to delete column: {str(e)}")
    
    def rename_lte_column(self):
        """Rename selected column in LTE treeview"""
        try:
            if hasattr(self, 'current_col_name'):
                new_name = simpledialog.askstring("Rename Column", "New column name:", initialvalue=self.current_col_name)
                if new_name and new_name != self.current_col_name:
                    current_columns = list(self.lte_tree['columns'])
                    if new_name in current_columns:
                        messagebox.showwarning("Duplicate", "Column name already exists")
                        return
                    idx = current_columns.index(self.current_col_name)
                    current_columns[idx] = new_name
                    self.rebuild_lte_tree(current_columns)
        except Exception as e:
            logging.error(f"Error in rename_lte_column: {str(e)}")
            messagebox.showerror("Error", f"Failed to rename column: {str(e)}")
    
    def generate_lte_cr(self):
        """Generate CR for selected LTE parameters"""
        try:
            selected = self.lte_tree.selection()
            if not selected:
                messagebox.showinfo("Info", "Please select at least one row")
                return
            
            cr_type = self.lte_cr_type.get()
            if not cr_type:
                messagebox.showinfo("Info", "Please select a CR type")
                return
            
            # Create DataFrame for CR output
            cr_data = []
            
            for item in selected:
                try:
                    values = self.lte_tree.item(item, 'values')
                    site = values[1]  # Site is at index 1
                    cell = values[2]  # cell is at index 2
                    
                    # Get parameter value based on CR type
                    if cr_type == "cellRange":
                        value = values[3]  # CELLRANGE is at index 3
                    elif cr_type == "crsGain":
                        value = values[4]  # CRSGAIN is at index 4
                    elif cr_type == "Electrical Tilt":
                        # Get from original record
                        record = self.lte_tree_record_map.get(item)
                        if record is not None:
                            tilt_cols = self.mappings["LTE"]["Electrical Tilt"]
                            value = self.get_column_value(record, tilt_cols, "LTE")
                        else:
                            value = ""
                    
                    # Create MO Class string
                    mo_class = f"EUtranCellFDD={cell}"
                    
                    # Add to CR data
                    cr_data.append({
                        "Site": site,
                        "MO Class": mo_class,
                        "Parameter": cr_type,
                        "Value": value,
                        "CurrentValue": ""  # Empty for engineer to fill
                    })
                except Exception as e:
                    logging.error(f"Error processing LTE row: {str(e)}")
                    continue
            
            if not cr_data:
                messagebox.showinfo("Info", "No valid data to generate CR")
                return
            
            # Create DataFrame
            df = pd.DataFrame(cr_data)
            
            # Save to Excel
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save CR File"
            )
            
            if not file_path:
                return
            
            df.to_excel(file_path, index=False)
            self.update_status(f"Generated {len(cr_data)} LTE CRs for {cr_type}")
            messagebox.showinfo("Success", "CR generated successfully!")
            
        except Exception as e:
            logging.error(f"Error in generate_lte_cr: {str(e)}")
            messagebox.showerror("Error", f"Failed to generate CR: {str(e)}")
    
    def create_5g_tab(self, parent):
        """Create the 5G Parameters tab with the exact layout from the screenshot"""
        results_frame = ttk.LabelFrame(parent, text="5G Parameters", padding=10)
        results_frame.pack(fill=tk.BOTH, expand=True)
        
        # Control frame for buttons and dropdowns
        control_frame = ttk.Frame(results_frame)
        control_frame.pack(fill=tk.X, pady=5)
        
        # Add Column button
        add_col_btn = ttk.Button(control_frame, text="Add Column", command=self.add_5g_column)
        add_col_btn.pack(side=tk.LEFT, padx=5)
        
        # CR Generation section
        ttk.Label(control_frame, text="Generate CR for:").pack(side=tk.LEFT, padx=5)
        self.nr_cr_type = ttk.Combobox(control_frame, 
                                      values=["digitalTilt", "cellRange"],
                                      width=15, state="readonly")
        self.nr_cr_type.current(0)
        self.nr_cr_type.pack(side=tk.LEFT, padx=5)
        
        # Generate 5GCR button
        cr_btn = ttk.Button(control_frame, text="Generate 5GCR", command=self.generate_5g_cr)
        cr_btn.pack(side=tk.LEFT, padx=5)
        
        # Export 5G Data to Excel button
        export_btn = ttk.Button(control_frame, text="Export 5G Data to Excel", 
                               command=self.export_5g_to_excel)
        export_btn.pack(side=tk.LEFT, padx=5)
        
        # Create a container frame for the treeview and scrollbars
        tree_container = ttk.Frame(results_frame)
        tree_container.pack(fill=tk.BOTH, expand=True)
        
        # Create Treeview with scrollbars
        columns = (
            "Source", "USID", "SITE", "NRCELL_NAME", "Digital Tilt", "Power", 
            "PCI", "ADMINISTRATIVESTATE", "CELLBARRED", "CELLRESERVEDFOROPERATOR", 
            "OPERATIONALSTATE", "CELLRANGE", "SSBFREQUENCY", "CONFIGURATION"
        )
        
        self.nr_tree = ttk.Treeview(tree_container, columns=columns, show="headings", selectmode="extended")
        vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self.nr_tree.yview)
        hsb = ttk.Scrollbar(tree_container, orient="horizontal", command=self.nr_tree.xview)
        self.nr_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Configure columns with headings
        col_widths = {
            "Source": 50, "USID": 70, "SITE": 90, "NRCELL_NAME": 110, 
            "Digital Tilt": 70, "Power": 50, "PCI": 40,
            "ADMINISTRATIVESTATE": 90, "CELLBARRED": 70, 
            "CELLRESERVEDFOROPERATOR": 90, "OPERATIONALSTATE": 90,
            "CELLRANGE": 70, "SSBFREQUENCY": 70, "CONFIGURATION": 70
        }
        
        for col in columns:
            self.nr_tree.heading(col, text=col)
            self.nr_tree.column(col, width=col_widths.get(col, 70), anchor=tk.CENTER)
        
        # Layout treeview and scrollbars
        self.nr_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        # Configure grid weights
        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)
        
        # Bind right-click
        self.nr_tree.bind("<Button-3>", self.handle_nr_right_click)
    
    def generate_5g_cr(self):
        """Generate 5G CR in the specified format"""
        try:
            selected = self.nr_tree.selection()
            if not selected:
                messagebox.showinfo("Info", "Please select at least one row")
                return
            
            cr_type = self.nr_cr_type.get()
            if not cr_type:
                messagebox.showinfo("Info", "Please select a CR type")
                return
            
            # Get current columns
            columns = self.nr_tree['columns']
            
            # Find indices of required columns
            site_idx = columns.index("SITE") if "SITE" in columns else -1
            nrcell_idx = columns.index("NRCELL_NAME") if "NRCELL_NAME" in columns else -1
            tilt_idx = columns.index("Digital Tilt") if "Digital Tilt" in columns else -1
            cellrange_idx = columns.index("CELLRANGE") if "CELLRANGE" in columns else -1
            
            if site_idx == -1 or nrcell_idx == -1:
                messagebox.showerror("Error", "Required columns (SITE, NRCELL_NAME) are not present in the 5G table")
                return
            
            # Create DataFrame for CR output
            cr_data = []
            
            for item in selected:
                try:
                    values = self.nr_tree.item(item, 'values')
                    site = values[site_idx]
                    nrcell = values[nrcell_idx]
                    
                    # Get parameter value based on CR type
                    if cr_type == "digitalTilt":
                        # Use digital tilt value if available
                        value = values[tilt_idx] if tilt_idx != -1 and tilt_idx < len(values) else ""
                        mo_class = f"NRSectorCarrier={nrcell},CommonBeamforming=1"
                    elif cr_type == "cellRange":
                        # Use cell range value if available
                        value = values[cellrange_idx] if cellrange_idx != -1 and cellrange_idx < len(values) else ""
                        mo_class = f"NRCellDU={nrcell}"
                    
                    # Add to CR data in the exact required format
                    cr_data.append({
                        "Site": site,
                        "MO Class": mo_class,
                        "Parameter": cr_type,
                        "Value": value,
                        "CurrentValue": ""  # Always empty
                    })
                except Exception as e:
                    logging.error(f"Error processing 5G row: {str(e)}")
                    continue
            
            if not cr_data:
                messagebox.showinfo("Info", "No valid data to generate CR")
                return
            
            # Create DataFrame
            df = pd.DataFrame(cr_data)
            
            # Save to Excel
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save 5G CR File"
            )
            
            if not file_path:
                return
            
            df.to_excel(file_path, index=False, sheet_name='5g cr buttun')
            self.update_status(f"Generated {len(cr_data)} 5G CRs for {cr_type}")
            messagebox.showinfo("Success", "5G CR generated successfully!")
            
        except Exception as e:
            logging.error(f"Error in generate_5g_cr: {str(e)}")
            messagebox.showerror("Error", f"Failed to generate 5G CR: {str(e)}")
    
    def create_vdt_tab(self, parent):
        """Create the VDT Sheet tab with updated layout"""
        vdt_frame = ttk.LabelFrame(parent, text="VDT Sheet", padding=10)
        vdt_frame.pack(fill=tk.BOTH, expand=True)
        
        # Project name selection and buttons in one row
        control_frame = ttk.Frame(vdt_frame)
        control_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(control_frame, text="Project Name:").pack(side=tk.LEFT, padx=5)
        project_combo = ttk.Combobox(control_frame, textvariable=self.project_name_var, 
                                    values=list(self.market_mapping.values()), state="readonly")
        project_combo.pack(side=tk.LEFT, padx=5)
        
        # Generate VDT Report button moved next to project name
        report_btn = ttk.Button(control_frame, text="Generate VDT Report", command=self.generate_vdt_report)
        report_btn.pack(side=tk.LEFT, padx=5)
        
        # Create a container frame for the treeview and scrollbars
        tree_container = ttk.Frame(vdt_frame)
        tree_container.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Create Treeview for VDT data
        columns = ("LTE Site", "NR Site")
        self.vdt_tree = ttk.Treeview(tree_container, columns=columns, show="headings", selectmode="extended")
        vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self.vdt_tree.yview)
        hsb = ttk.Scrollbar(tree_container, orient="horizontal", command=self.vdt_tree.xview)
        self.vdt_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Configure columns
        for col in columns:
            self.vdt_tree.heading(col, text=col)
            self.vdt_tree.column(col, width=150, anchor=tk.CENTER)
        
        # Layout treeview and scrollbars using grid within the container
        self.vdt_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        # Configure grid weights for the container
        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)
        
        # Add auto-generate toggle
        auto_frame = ttk.Frame(vdt_frame)
        auto_frame.pack(fill=tk.X, pady=5)
        
        auto_cb = ttk.Checkbutton(auto_frame, text="Auto-Generate on Search", 
                                 variable=self.auto_generate_var)
        auto_cb.pack(side=tk.LEFT, padx=5)
    
    def generate_vdt_data(self):
        """Generate VDT data from unique LTE and 5G sites in parameter tabs"""
        try:
            # Clear existing data
            for item in self.vdt_tree.get_children():
                self.vdt_tree.delete(item)
            
            # Get unique LTE sites from LTE parameter tab
            lte_sites = set()
            for item in self.lte_tree.get_children():
                values = self.lte_tree.item(item, 'values')
                if len(values) > 1:  # Ensure we have Site column
                    site = values[1]
                    if site:
                        lte_sites.add(site)
            
            # Get unique 5G sites from 5G parameter tab
            nr_sites = set()
            columns = self.nr_tree['columns']
            site_idx = columns.index("SITE") if "SITE" in columns else -1
            if site_idx != -1:
                for item in self.nr_tree.get_children():
                    values = self.nr_tree.item(item, 'values')
                    if len(values) > site_idx:
                        site = values[site_idx]
                        if site:
                            nr_sites.add(site)
            
            # Convert to sorted lists
            lte_sites = sorted(lte_sites)
            nr_sites = sorted(nr_sites)
            
            # Create pairs (pad with empty if needed)
            max_len = max(len(lte_sites), len(nr_sites))
            for i in range(max_len):
                lte_site = lte_sites[i] if i < len(lte_sites) else ""
                nr_site = nr_sites[i] if i < len(nr_sites) else ""
                self.vdt_tree.insert("", "end", values=(lte_site, nr_site))
            
            self.update_status(f"Generated VDT data with {max_len} rows")
            
        except Exception as e:
            logging.error(f"Error in generate_vdt_data: {str(e)}")
            messagebox.showerror("Error", f"Failed to generate VDT data: {str(e)}")
    
    def generate_vdt_report(self):
        """Generate VDT report in the specified Excel format"""
        try:
            # Get unique sites from VDT tab
            lte_sites = []
            nr_sites = []
            
            for item in self.vdt_tree.get_children():
                values = self.vdt_tree.item(item, 'values')
                if values[0]:  # LTE Site
                    lte_sites.append(values[0])
                if values[1]:  # NR Site
                    nr_sites.append(values[1])
            
            if not lte_sites and not nr_sites:
                messagebox.showinfo("Info", "No sites found in VDT tab. Please generate VDT data first.")
                return
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create LTE sheet
            lte_sheet = wb.create_sheet("LTE")
            
            # Add headers
            lte_sheet.append(["Type", "Value", "Remarks", "Buffer", ""])
            
            # Add fixed content (light blue)
            project_name = self.project_name_var.get()
            lte_sheet.append(["projectName", project_name, "", "0.01", ""])
            
            # Add current time (gray)
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            lte_sheet.append(["startTime", current_time, "Use exact format. Add \" ' \" before", "Default keep it 0.01. Can increase it to accommodate TA Plot", ""])
            lte_sheet.append(["endTime", current_time, "Use exact format. Add \" ' \" before", "", ""])
            
            # Add site list header
            lte_sheet.append(["Site Name List", "ZOOM (Default = 0 for automatic zoom, enter integer value for manual zoom\nEg. value: 5 will fetch plots up to 5 km diagonally from site in both directions.", 
                             "Neighbour list (comma separated), if blank first tier neighbours will be considered", 
                             "Carrier List(comma separated)", "EARFCN DL from lte parameter data source"])
            
            # Add sites (orange) and EARFCN values (yellow)
            for site in lte_sites:
                # Get EARFCN values for this site from LTE parameter tab
                earfcn_values = set()
                for item in self.lte_tree.get_children():
                    values = self.lte_tree.item(item, 'values')
                    if len(values) > 1 and values[1] == site:  # Site matches
                        # Get EARFCNDL value if available
                        earfcn_idx = self.lte_tree['columns'].index("EARFCNDL") if "EARFCNDL" in self.lte_tree['columns'] else -1
                        if earfcn_idx != -1 and earfcn_idx < len(values):
                            earfcn = values[earfcn_idx]
                            if earfcn:
                                earfcn_values.add(earfcn)
                
                # Join EARFCN values with commas
                earfcn_str = ",".join(sorted(earfcn_values)) if earfcn_values else ""
                
                lte_sheet.append([site, "0", "", earfcn_str, ""])
            
            # Create NR sheet
            nr_sheet = wb.create_sheet("NR")
            
            # Add headers
            nr_sheet.append(["Type", "Value", "Remarks", "Buffer", "", ""])
            
            # Add fixed content (light blue)
            nr_sheet.append(["projectName", project_name, "", "0.01", "", ""])
            
            # Add current time (gray)
            nr_sheet.append(["startTime", current_time, "Use exact format. Add \" ' \" before", "Default keep it 0.01. Can increase it to accommodate TA Plot", "", ""])
            nr_sheet.append(["endTime", current_time, "Use exact format. Add \" ' \" before", "", "", ""])
            
            # Add site list header
            nr_sheet.append(["Site Name List", "ZOOM (Default = 0 for automatic zoom, enter integer value for manual zoom\nEg. value: 5 will fetch plots up to 5 km diagonally from site in both directions.", 
                            "Neighbour list (comma separated), if blank first tier neighbours will be considered", 
                            "Carrier List(comma separated)", "ARFCNDL", "from data source"])
            
            # Add sites (orange) and ARFCNDL values (yellow)
            for site in nr_sites:
                # Get ARFCNDL values for this site from 5G parameter tab
                arfcn_values = set()
                columns = self.nr_tree['columns']
                site_idx = columns.index("SITE") if "SITE" in columns else -1
                arfcn_idx = columns.index("SSBFREQUENCY") if "SSBFREQUENCY" in columns else -1
                
                if site_idx != -1 and arfcn_idx != -1:
                    for item in self.nr_tree.get_children():
                        values = self.nr_tree.item(item, 'values')
                        if len(values) > site_idx and values[site_idx] == site:  # Site matches
                            if arfcn_idx < len(values):
                                arfcn = values[arfcn_idx]
                                if arfcn:
                                    arfcn_values.add(arfcn)
                
                # Join ARFCN values with commas
                arfcn_str = ",".join(sorted(arfcn_values)) if arfcn_values else ""
                
                nr_sheet.append([site, "0", "", arfcn_str, "ARFCNDL", "from data source"])
            
            # Save file
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save VDT Report"
            )
            
            if not file_path:
                return
            
            wb.save(file_path)
            self.update_status(f"VDT report generated with {len(lte_sites)} LTE and {len(nr_sites)} NR sites")
            messagebox.showinfo("Success", "VDT report generated successfully!")
            
        except Exception as e:
            logging.error(f"Error in generate_vdt_report: {str(e)}")
            messagebox.showerror("Error", f"Failed to generate VDT report: {str(e)}")
            
    def create_distance_tab(self, parent):
        """Create the enhanced distance calculator tab"""
        dist_frame = ttk.Frame(parent)
        dist_frame.pack(fill=tk.BOTH, expand=True)
        
        # Split into left (controls) and right (results/map) panes
        paned_window = ttk.PanedWindow(dist_frame, orient=tk.HORIZONTAL)
        paned_window.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Left pane - controls
        control_frame = ttk.LabelFrame(paned_window, text="Distance Calculator", padding=10)
        paned_window.add(control_frame, weight=1)
        
        # Point input with separate lat/lon
        input_frame = ttk.Frame(control_frame)
        input_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(input_frame, text="Point Name:").pack(side=tk.LEFT, padx=5)
        self.point_name_entry = ttk.Entry(input_frame, width=15)
        self.point_name_entry.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(input_frame, text="Latitude:").pack(side=tk.LEFT, padx=5)
        self.lat_entry = ttk.Entry(input_frame, width=12)
        self.lat_entry.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(input_frame, text="Longitude:").pack(side=tk.LEFT, padx=5)
        self.lon_entry = ttk.Entry(input_frame, width=12)
        self.lon_entry.pack(side=tk.LEFT, padx=5)
        
        add_btn = ttk.Button(input_frame, text="Add", command=self.add_point)
        add_btn.pack(side=tk.LEFT, padx=5)
        
        # New lat/long input box
        lat_long_frame = ttk.Frame(control_frame)
        lat_long_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(lat_long_frame, text="Lat,Long (e.g., 12.34,78.56):").pack(side=tk.LEFT, padx=5)
        self.lat_long_entry = ttk.Entry(lat_long_frame, width=20)
        self.lat_long_entry.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(lat_long_frame, text="Point Name:").pack(side=tk.LEFT, padx=5)
        self.lat_long_name_entry = ttk.Entry(lat_long_frame, width=15)
        self.lat_long_name_entry.pack(side=tk.LEFT, padx=5)
        
        add_lat_long_btn = ttk.Button(lat_long_frame, text="Add", command=self.add_point_from_lat_long)
        add_lat_long_btn.pack(side=tk.LEFT, padx=5)
        
        # Excel import
        import_frame = ttk.Frame(control_frame)
        import_frame.pack(fill=tk.X, pady=5)
        
        import_btn = ttk.Button(import_frame, text="Import from Excel", command=self.import_from_excel)
        import_btn.pack(side=tk.LEFT, padx=5)
        
        # Calculate buttons
        calc_frame = ttk.Frame(control_frame)
        calc_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(calc_frame, text="Calculate Path", command=self.calculate_path_distances).pack(side=tk.LEFT, padx=5)
        ttk.Button(calc_frame, text="Calculate from Master", command=self.calculate_from_master).pack(side=tk.LEFT, padx=5)
        ttk.Button(calc_frame, text="Show on Google Maps", command=self.open_google_maps).pack(side=tk.LEFT, padx=5)
        
        # Points list
        list_frame = ttk.Frame(control_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        ttk.Label(list_frame, text="Points List:").pack(anchor="w")
        self.points_listbox = tk.Listbox(list_frame, height=8)
        self.points_listbox.pack(fill=tk.BOTH, expand=True, pady=5)
        # Bind left-click to prompt for master point
        self.points_listbox.bind("<Button-1>", self.prompt_master_point)
        
        # Buttons for point management
        btn_frame = ttk.Frame(control_frame)
        btn_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(btn_frame, text="Move Up", command=lambda: self.move_point(-1)).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="Move Down", command=lambda: self.move_point(1)).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="Set as Master", command=self.set_as_master).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="Remove", command=self.remove_point).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="Clear All", command=self.clear_points).pack(side=tk.LEFT, padx=2)
        
        # Right pane - results and map
        result_frame = ttk.LabelFrame(paned_window, text="Results & Map", padding=10)
        paned_window.add(result_frame, weight=2)
        
        # Create notebook for results and map
        self.dist_notebook = ttk.Notebook(result_frame)
        self.dist_notebook.pack(fill=tk.BOTH, expand=True)
        
        # Results tab
        results_tab = ttk.Frame(self.dist_notebook)
        self.dist_notebook.add(results_tab, text="Results")
        
        self.distance_text = scrolledtext.ScrolledText(results_tab)
        self.distance_text.pack(fill=tk.BOTH, expand=True)
        self.distance_text.config(state=tk.DISABLED)
        
        # Configure text tags for formatting
        self.distance_text.tag_configure("header", font=("Arial", 10, "bold"))
        self.distance_text.tag_configure("bold", font=("Arial", 9, "bold"))
        
        # Map tab
        map_tab = ttk.Frame(self.dist_notebook)
        self.dist_notebook.add(map_tab, text="Map Visualization")
        
        # Map display
        map_control_frame = ttk.Frame(map_tab)
        map_control_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(map_control_frame, text="Map Type:").pack(side=tk.LEFT, padx=5)
        self.map_type = ttk.Combobox(map_control_frame, 
                                   values=["roadmap", "satellite", "hybrid", "terrain"],
                                   width=10, state="readonly")
        self.map_type.current(0)
        self.map_type.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(map_control_frame, text="Zoom:").pack(side=tk.LEFT, padx=5)
        self.map_zoom = ttk.Spinbox(map_control_frame, from_=1, to=20, width=5)
        self.map_zoom.set(12)
        self.map_zoom.pack(side=tk.LEFT, padx=5)
        
        show_map_btn = ttk.Button(map_control_frame, text="Show Map", command=self.show_map)
        show_map_btn.pack(side=tk.RIGHT, padx=5)
        
        # Map image display
        self.map_label = ttk.Label(map_tab)
        self.map_label.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Add instructions
        ttk.Label(control_frame, text="Tip: Right-click on search results to add points").pack(anchor="w")
    
    def browse_files(self):
        """Open file dialog to select multiple data files"""
        try:
            file_paths = filedialog.askopenfilenames(
                title="Select Network Data Files",
                filetypes=[("CSV Files", "*.csv"), ("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")]
            )
            
            if file_paths:
                self.file_listbox.delete(0, tk.END)
                for path in file_paths:
                    self.file_listbox.insert(tk.END, path)
        except Exception as e:
            logging.error(f"Error in browse_files: {str(e)}")
            messagebox.showerror("Error", f"Failed to browse files: {str(e)}")
    
    def load_data(self):
        """Load data from selected files"""
        try:
            file_paths = self.file_listbox.get(0, tk.END)
            
            if not file_paths:
                messagebox.showwarning("Input Error", "Please select at least one data file")
                return
            
            self.lte_data = pd.DataFrame()
            self.nr_data = pd.DataFrame()
            self.bbu_data = pd.DataFrame()
            self.usid_index = {}
            self.bbu_index = {}
            
            for file_path in file_paths:
                self.update_status(f"Loading data from {os.path.basename(file_path)}")
                
                # Read file
                if file_path.endswith('.csv'):
                    df = pd.read_csv(file_path, dtype=str)
                else:
                    df = pd.read_excel(file_path, dtype=str)
                
                # Determine file type with more robustness
                filename = os.path.basename(file_path).upper()
                columns = df.columns.str.upper().tolist()
                
                if 'LTE' in filename or any(col in columns for col in ['ENBID', 'CELLID', 'EUTRAN_CELL_FDD_ID']):
                    # Remove duplicates before concatenating
                    df = df.drop_duplicates()
                    self.lte_data = pd.concat([self.lte_data, df], ignore_index=True)
                elif 'BBU' in filename or any(col in columns for col in ['BBU_TECH', 'GNB_SA_STATE']):
                    # Remove duplicates before concatenating
                    df = df.drop_duplicates()
                    self.bbu_data = pd.concat([self.bbu_data, df], ignore_index=True)
                elif '5G' in filename or 'NR' in filename or any(col in columns for col in ['NCI', 'GNBID', 'NRCELLDU', 'NRCELLDUID']):
                    # Remove duplicates before concatenating
                    df = df.drop_duplicates()
                    self.nr_data = pd.concat([self.nr_data, df], ignore_index=True)
                else:
                    # Fallback: check for key keywords
                    lte_keywords = ['ENBID', 'CELLID', 'EUTRAN']
                    nr_keywords = ['NCI', 'GNBID', 'NRCELL']
                    bbu_keywords = ['BBU_TECH', 'GNB_SA']
                    scores = {
                        'LTE': sum(1 for kw in lte_keywords if any(kw in col for col in columns)),
                        '5GNR': sum(1 for kw in nr_keywords if any(kw in col for col in columns)),
                        '5GNR_BBU': sum(1 for kw in bbu_keywords if any(kw in col for col in columns))
                    }
                    max_score_type = max(scores, key=scores.get)
                    if scores[max_score_type] > 0:
                        # Remove duplicates before concatenating
                        df = df.drop_duplicates()
                        if max_score_type == 'LTE':
                            self.lte_data = pd.concat([self.lte_data, df], ignore_index=True)
                        elif max_score_type == '5GNR':
                            self.nr_data = pd.concat([self.nr_data, df], ignore_index=True)
                        elif max_score_type == '5GNR_BBU':
                            self.bbu_data = pd.concat([self.bbu_data, df], ignore_index=True)
            
            # Remove duplicates from the final dataframes
            self.lte_data = self.lte_data.drop_duplicates()
            self.nr_data = self.nr_data.drop_duplicates()
            self.bbu_data = self.bbu_data.drop_duplicates()
            
            # Build USID index
            self.build_index()
            
            # Build BBU index
            self.build_bbu_index()
            
            messagebox.showinfo("Success", "Data loaded successfully!")
            self.update_status(f"Loaded {len(self.lte_data)} LTE, {len(self.nr_data)} 5GNR, and {len(self.bbu_data)} BBU records")
        
        except Exception as e:
            logging.error(f"Error in load_data: {str(e)}")
            self.update_status(f"Error loading data: {str(e)}")
            messagebox.showerror("Error", f"Failed to load data: {str(e)}")
    
    def clean_value(self, value):
        """Clean numeric values to remove .0 suffix"""
        try:
            if pd.isna(value) or value == "nan" or value is None:
                return ""
            # Remove decimal part if it's .0
            if isinstance(value, float) and value.is_integer():
                return str(int(value))
            if isinstance(value, str) and '.' in value and value.split('.')[1] == '0':
                return value.split('.')[0]
            return str(value)
        except:
            return str(value)
    
    def build_index(self):
        """Build USID index for fast searching"""
        try:
            self.usid_index = {}
            
            # Index LTE data
            if not self.lte_data.empty:
                usid_cols = self.mappings["LTE"]["USID"]
                for _, row in self.lte_data.iterrows():
                    usid = self.get_column_value(row, usid_cols, "LTE")
                    if usid:
                        # Store the entire row
                        self.usid_index.setdefault(usid, []).append(("LTE", row))
            
            # Index 5GNR data
            if not self.nr_data.empty:
                usid_cols = self.mappings["5GNR"]["USID"]
                for _, row in self.nr_data.iterrows():
                    usid = self.get_column_value(row, usid_cols, "5GNR")
                    if usid:
                        # Store the entire row
                        self.usid_index.setdefault(usid, []).append(("5GNR", row))
            
            self.update_status(f"Indexed {len(self.usid_index)} unique USIDs")
        except Exception as e:
            logging.error(f"Error in build_index: {str(e)}")
            self.update_status("Error building USID index")
    
    def build_bbu_index(self):
        """Build index for BBU data based on USID and NRCELL_NAME"""
        try:
            self.bbu_index = {}
            
            if not self.bbu_data.empty:
                usid_cols = self.mappings["5GNR_BBU"]["USID"]
                nrcell_cols = self.mappings["5GNR_BBU"]["NRCELL_NAME"]
                
                for _, row in self.bbu_data.iterrows():
                    usid = self.get_column_value(row, usid_cols, "5GNR_BBU")
                    nrcell = self.get_column_value(row, nrcell_cols, "5GNR_BBU")
                    
                    if usid and nrcell:
                        # Create composite key
                        key = (usid, nrcell)
                        self.bbu_index[key] = row
            
            self.update_status(f"Indexed {len(self.bbu_index)} BBU records")
        except Exception as e:
            logging.error(f"Error in build_bbu_index: {str(e)}")
            self.update_status("Error building BBU index")
    
    def perform_search(self):
        """Execute search based on user input"""
        try:
            # Clear previous results
            self.clear_results()
            
            search_type = self.search_type.get()
            search_value = self.search_entry.get().strip()
            
            if not search_value:
                messagebox.showwarning("Input Error", "Please enter a search value")
                return
            
            # Find specific matching records
            matched_records = self.find_matching_records(search_type, search_value)
            
            if not matched_records:
                messagebox.showinfo("No Results", "No matching records found")
                self.update_status("Search completed with no results")
                return
            
            # Merge records to ensure uniqueness and fill missing values
            merged_records = self.merge_records(matched_records)
            
            # Store matched records for all tabs
            self.matched_records = merged_records
            
            # Add the matched records to results
            for tech, record in merged_records:
                self.add_to_main_tree(tech, record)
                self.add_to_lte_tree(tech, record)
                self.add_to_5g_tree(tech, record)
            
            # Auto-generate VDT data if enabled
            if self.auto_generate_var.get():
                self.generate_vdt_data()
            
            self.update_status(f"Found {len(merged_records)} records for {search_type}={search_value}")
        except Exception as e:
            logging.error(f"Error in perform_search: {str(e)}")
            messagebox.showerror("Search Error", f"Search failed: {str(e)}")
    
    def merge_records(self, records):
        """Merge duplicate records by filling missing values"""
        from collections import defaultdict
        merged = []
        lte_groups = defaultdict(list)
        nr_groups = defaultdict(list)

        for tech, row in records:
            if tech == "LTE":
                cell = self.get_column_value(row, self.mappings["LTE"]["cell"], "LTE")
                lte_groups[cell].append(row)
            elif tech == "5GNR":
                cell = self.get_column_value(row, self.mappings["5GNR"]["cell"], "5GNR")
                nr_groups[cell].append(row)

        # Merge LTE
        for cell, rows in lte_groups.items():
            if not rows:
                continue
            # Merge: first row as base, fill from others
            base = rows[0].copy()
            for other in rows[1:]:
                for col in base.index:
                    if pd.isna(base[col]) or base[col] == "":
                        base[col] = other.get(col, base[col])
            merged.append(("LTE", base))

        # Merge NR
        for cell, rows in nr_groups.items():
            if not rows:
                continue
            base = rows[0].copy()
            for other in rows[1:]:
                for col in base.index:
                    if pd.isna(base[col]) or base[col] == "":
                        base[col] = other.get(col, base[col])
            merged.append(("5GNR", base))

        return merged
    
    def find_matching_records(self, search_type, value):
        """Find specific matching records based on search type and value"""
        try:
            matched_records = []
            value = self.clean_value(value)
            
            # Search through LTE data
            if search_type in ["cell ID", "ENBID", "Site"] and not self.lte_data.empty:
                lte_mapping = self.mappings["LTE"]
                
                # Search by cell ID
                if search_type == "cell ID":
                    col_names = lte_mapping["cell ID"]
                    for _, row in self.lte_data.iterrows():
                        cell_id = self.get_column_value(row, col_names, "LTE")
                        if cell_id == value:
                            matched_records.append(("LTE", row))
                
                # Search by ENBID
                elif search_type == "ENBID":
                    col_names = lte_mapping["ENBID"]
                    for _, row in self.lte_data.iterrows():
                        enbid = self.get_column_value(row, col_names, "LTE")
                        if enbid == value:
                            matched_records.append(("LTE", row))
                
                # Search by Site
                elif search_type == "Site":
                    col_names = lte_mapping["Site"]
                    for _, row in self.lte_data.iterrows():
                        site = self.get_column_value(row, col_names, "LTE")
                        if site == value:
                            matched_records.append(("LTE", row))
            
            # Search through 5GNR data
            elif search_type in ["NIC", "gnb ID", "Site"] and not self.nr_data.empty:
                nr_mapping = self.mappings["5GNR"]
                
                # Search by NIC
                if search_type == "NIC":
                    col_names = nr_mapping["NIC"]
                    for _, row in self.nr_data.iterrows():
                        nic = self.get_column_value(row, col_names, "5GNR")
                        if nic == value:
                            matched_records.append(("5GNR", row))
                
                # Search by gnb ID
                elif search_type == "gnb ID":
                    col_names = nr_mapping["gnb ID"]
                    for _, row in self.nr_data.iterrows():
                        gnb_id = self.get_column_value(row, col_names, "5GNR")
                        if gnb_id == value:
                            matched_records.append(("5GNR", row))
                
                # Search by Site
                elif search_type == "Site":
                    col_names = nr_mapping["Site"]
                    for _, row in self.nr_data.iterrows():
                        site = self.get_column_value(row, col_names, "5GNR")
                        if site == value:
                            matched_records.append(("5GNR", row))
            
            # Direct USID search
            elif search_type == "USID":
                if value in self.usid_index:
                    matched_records = self.usid_index[value]
            
            return matched_records
        except Exception as e:
            logging.error(f"Error in find_matching_records: {str(e)}")
            return []
    
    def get_column_value(self, record, possible_names, tech):
        """Robust column value extraction with comprehensive matching and fuzzy"""
        try:
            # First try exact matches
            for name in possible_names:
                if name in record:
                    value = record[name]
                    if not pd.isna(value) and value != "":
                        return self.clean_value(value)
            
            # Try case-insensitive match
            for col in record.index:
                col_str = str(col).strip().lower()
                for name in possible_names:
                    if col_str == name.lower():
                        value = record[col]
                        if not pd.isna(value) and value != "":
                            return self.clean_value(value)
            
            # Try substring match
            for col in record.index:
                col_str = str(col).strip().lower()
                for name in possible_names:
                    if name.lower() in col_str:
                        value = record[col]
                        if not pd.isna(value) and value != "":
                            return self.clean_value(value)
            
            # Fuzzy matching using difflib
            all_cols = list(record.index)
            for name in possible_names:
                close_matches = difflib.get_close_matches(name.lower(), [c.lower() for c in all_cols], n=1, cutoff=0.8)
                if close_matches:
                    matched_col = all_cols[[c.lower() for c in all_cols].index(close_matches[0])]
                    value = record[matched_col]
                    if not pd.isna(value) and value != "":
                        return self.clean_value(value)
            
            # Special handling for 5GNR cell values
            if tech == "5GNR" and "NRCELL_NAME" in possible_names:
                for col in record.index:
                    col_str = str(col).strip().lower()
                    if "nrcell" in col_str or "cell" in col_str or "name" in col_str:
                        value = record[col]
                        if not pd.isna(value) and value != "":
                            return self.clean_value(value)
            
            # Fill from BBU for 5GNR if applicable
            if tech == "5GNR" and possible_names and possible_names[0] in self.mappings["5GNR_BBU"]:
                usid = self.get_column_value(record, self.mappings["5GNR"]["USID"], "5GNR")
                nrcell = self.get_column_value(record, self.mappings["5GNR"]["cell"], "5GNR")
                if usid and nrcell:
                    key = (usid, nrcell)
                    if key in self.bbu_index:
                        bbu_row = self.bbu_index[key]
                        bbu_possible = self.mappings["5GNR_BBU"].get(possible_names[0], possible_names)
                        bbu_value = ""
                        for name in bbu_possible:
                            if name in bbu_row:
                                bbu_value = bbu_row[name]
                                if not pd.isna(bbu_value) and bbu_value != "":
                                    return self.clean_value(bbu_value)
                        
                        # Case insensitive for BBU
                        for col in bbu_row.index:
                            col_str = str(col).strip().lower()
                            for name in bbu_possible:
                                if col_str == name.lower():
                                    bbu_value = bbu_row[col]
                                    if not pd.isna(bbu_value) and bbu_value != "":
                                        return self.clean_value(bbu_value)
            
            return ""
        except Exception as e:
            logging.error(f"Error in get_column_value: {str(e)}")
            return ""
    
    def add_to_main_tree(self, tech, record):
        """Add a record to the main results treeview"""
        try:
            # Get the appropriate mapping for the technology
            mapping = self.mappings[tech]
            
            # Extract values using the mapping with multiple possible names
            values = {
                "Source": tech,
                "NIC": self.get_column_value(record, mapping.get("NIC", []), tech),
                "gnb ID": self.get_column_value(record, mapping.get("gnb ID", []), tech),
                "ENBID": self.get_column_value(record, mapping.get("ENBID", []), tech),
                "cell ID": self.get_column_value(record, mapping.get("cell ID", []), tech),
                "USID": self.get_column_value(record, mapping.get("USID", []), tech),
                "Site": self.get_column_value(record, mapping.get("Site", []), tech),
                "Azumuth": self.get_column_value(record, mapping.get("Azumuth", []), tech),
                "Digital Tilt": self.get_column_value(record, mapping.get("Digital Tilt", []), tech),
                "cell": self.get_column_value(record, mapping.get("cell", []), tech),
                "height(Meter)": self.get_column_value(record, mapping.get("height(Meter)", []), tech),
                "PCI": self.get_column_value(record, mapping.get("PCI", []), tech),
                "Power": self.get_column_value(record, mapping.get("Power", []), tech),
                "LATITUDE": self.get_column_value(record, mapping.get("LATITUDE", []), tech),
                "LONGITUDE": self.get_column_value(record, mapping.get("LONGITUDE", []), tech),
                "ADMINISTRATIVESTATE": self.get_column_value(record, mapping.get("ADMINISTRATIVESTATE", []), tech),
                "OPERATIONALSTATE": self.get_column_value(record, mapping.get("OPERATIONALSTATE", []), tech)
            }
            
            # Insert into treeview
            item = self.tree.insert("", "end", values=(
                values["Source"],
                values["NIC"],
                values["gnb ID"],
                values["ENBID"],
                values["cell ID"],
                values["USID"],
                values["Site"],
                values["Azumuth"],
                values["Digital Tilt"],
                values["cell"],
                values["height(Meter)"],
                values["PCI"],
                values["Power"],
                values["LATITUDE"],
                values["LONGITUDE"],
                values["ADMINISTRATIVESTATE"],
                values["OPERATIONALSTATE"]
            ))
            
            # Highlight search-related columns
            highlight_cols = ["NIC", "gnb ID", "ENBID", "cell ID", "USID", "Site"]
            for col in highlight_cols:
                if values[col]:
                    self.tree.item(item, tags=('highlight',))
        except Exception as e:
            logging.error(f"Error in add_to_main_tree: {str(e)}")
    
    def add_to_lte_tree(self, tech, record):
        """Add LTE parameters to the LTE tab treeview with added Site column"""
        try:
            # Only show LTE records in this tab
            if tech != "LTE":
                return
                
            mapping = self.mappings["LTE"]
            
            # Extract values for LTE parameters
            values = {
                "Source": tech,
                "Site": self.get_column_value(record, mapping.get("Site", []), tech),
                "cell": self.get_column_value(record, mapping.get("cell", []), tech),
                "CELLRANGE": self.get_column_value(record, mapping.get("CELLRANGE", []), tech),
                "CRSGAIN": self.get_column_value(record, mapping.get("CRSGAIN", []), tech),
                "QRXLEVMIN": self.get_column_value(record, mapping.get("QRXLEVMIN", []), tech),
                "EARFCNDL": self.get_column_value(record, mapping.get("EARFCNDL", []), tech)
            }
            
            # Insert into LTE treeview
            item = self.lte_tree.insert("", "end", values=(
                values["Source"],
                values["Site"],
                values["cell"],
                values["CELLRANGE"],
                values["CRSGAIN"],
                values["QRXLEVMIN"],
                values["EARFCNDL"]
            ))
            
            # Store record reference for VDT and CR
            self.lte_tree_record_map[item] = record
            
            # Highlight the cell column
            if values["cell"]:
                self.lte_tree.item(item, tags=('highlight',))
        except Exception as e:
            logging.error(f"Error in add_to_lte_tree: {str(e)}")
    
    def add_to_5g_tree(self, tech, record):
        """Add 5G records to the 5G tab with dynamic columns"""
        try:
            if tech != "5GNR":
                return
                
            # Get current columns
            columns = self.nr_tree['columns']
            
            # Column to mapping key translation
            column_to_key = {
                "USID": "USID",
                "SITE": "Site",
                "NRCELL_NAME": "cell",
                "Digital Tilt": "Digital Tilt",
                "Power": "Power",
                "PCI": "PCI",
                "ADMINISTRATIVESTATE": "ADMINISTRATIVESTATE",
                "CELLBARRED": "CELLBARRED",
                "CELLRESERVEDFOROPERATOR": "CELLRESERVEDFOROPERATOR",
                "OPERATIONALSTATE": "OPERATIONALSTATE",
                "CELLRANGE": "CELLRANGE",
                "SSBFREQUENCY": "SSBFREQUENCY",
                "ARFCNDL": "ARFCNDL",
                "CONFIGURATION": "CONFIGURATION",
                "ED_Market": "ED_Market"
            }
            
            # Prepare the list of values for the row
            values = []
            for col in columns:
                if col == "Source":
                    values.append(tech)
                else:
                    key = column_to_key.get(col, col)
                    if key in self.mappings["5GNR"]:
                        candidate_columns = self.mappings["5GNR"][key]
                    else:
                        candidate_columns = [col]
                    value = self.get_column_value(record, candidate_columns, "5GNR")
                    values.append(value)
            
            # Insert the row
            item = self.nr_tree.insert("", "end", values=values)
            
            # Store record reference for VDT and CR
            self.nr_tree_record_map[item] = record
            
            # Highlight if we have USID or NRCELL_NAME
            highlight = False
            if "USID" in columns:
                usid_index = columns.index("USID")
                if values[usid_index]:
                    highlight = True
            if not highlight and "NRCELL_NAME" in columns:
                nrcell_index = columns.index("NRCELL_NAME")
                if values[nrcell_index]:
                    highlight = True
            
            if highlight:
                self.nr_tree.item(item, tags=('highlight',))
        except Exception as e:
            logging.error(f"Error in add_to_5g_tree: {str(e)}")
    
    def show_context_menu(self, event):
        """Show right-click context menu"""
        try:
            selection = self.tree.selection()
            if not selection:
                self.context_menu.entryconfig("Copy with Headers", state=tk.DISABLED)
                self.context_menu.entryconfig("Export Selected to Excel", state=tk.DISABLED)
                self.context_menu.entryconfig("Use for Distance Calculation", state=tk.DISABLED)
            else:
                item = selection[0]
                values = self.tree.item(item, 'values')
                col_names = self.tree['columns']
                if "LATITUDE" in col_names and "LONGITUDE" in col_names:
                    try:
                        lat_idx = col_names.index("LATITUDE")
                        lon_idx = col_names.index("LONGITUDE")
                        lat_str = values[lat_idx]
                        lon_str = values[lon_idx]
                        if lat_str and lon_str:
                            float(lat_str)
                            float(lon_str)
                            self.context_menu.entryconfig("Use for Distance Calculation", state=tk.NORMAL)
                        else:
                            self.context_menu.entryconfig("Use for Distance Calculation", state=tk.DISABLED)
                    except (ValueError, IndexError):
                        self.context_menu.entryconfig("Use for Distance Calculation", state=tk.DISABLED)
                else:
                    self.context_menu.entryconfig("Use for Distance Calculation", state=tk.DISABLED)
                
                # Enable other commands
                self.context_menu.entryconfig("Copy with Headers", state=tk.NORMAL)
                self.context_menu.entryconfig("Export Selected to Excel", state=tk.NORMAL)
                
            self.context_menu.entryconfig("Clear Results", state=tk.NORMAL)
            self.context_menu.tk_popup(event.x_root, event.y_root)
        except Exception as e:
            logging.error(f"Error in show_context_menu: {str(e)}")
    
    def clear_results(self):
        """Clear all results from the treeviews"""
        try:
            for item in self.tree.get_children():
                self.tree.delete(item)
            for item in self.lte_tree.get_children():
                self.lte_tree.delete(item)
            for item in self.nr_tree.get_children():  # Clear 5G tab as well
                self.nr_tree.delete(item)
            for item in self.vdt_tree.get_children():  # Clear VDT tab
                self.vdt_tree.delete(item)
            
            # Clear record maps
            self.lte_tree_record_map = {}
            self.nr_tree_record_map = {}
        except Exception as e:
            logging.error(f"Error in clear_results: {str(e)}")
    
    def copy_with_headers(self):
        """Copy selected rows with headers to clipboard"""
        try:
            selected = self.tree.selection()
            if not selected:
                return
            
            # Get headers
            headers = self.tree['columns']
            
            # Create header row
            header_text = "\t".join(headers)
            
            # Create data rows
            data_rows = []
            for item in selected:
                values = self.tree.item(item, 'values')
                data_rows.append("\t".join(values))
            
            # Combine header and data
            clipboard_text = header_text + "\n" + "\n".join(data_rows)
            
            # Copy to clipboard
            self.root.clipboard_clear()
            self.root.clipboard_append(clipboard_text)
            self.update_status("Copied selected rows with headers to clipboard")
        except Exception as e:
            logging.error(f"Error in copy_with_headers: {str(e)}")
            messagebox.showerror("Error", f"Failed to copy: {str(e)}")
    
    def export_to_excel(self):
        """Export all results to Excel"""
        try:
            self.export_results(self.tree.get_children())
        except Exception as e:
            logging.error(f"Error in export_to_excel: {str(e)}")
            messagebox.showerror("Error", f"Export failed: {str(e)}")
    
    def export_selected_to_excel(self):
        """Export selected results to Excel"""
        try:
            selected = self.tree.selection()
            if not selected:
                messagebox.showinfo("Info", "No rows selected for export")
                return
            self.export_results(selected)
        except Exception as e:
            logging.error(f"Error in export_selected_to_excel: {str(e)}")
            messagebox.showerror("Error", f"Export failed: {str(e)}")
    
    def export_lte_to_excel(self):
        """Export LTE parameters to Excel"""
        try:
            items = self.lte_tree.get_children()
            if not items:
                messagebox.showinfo("Info", "No LTE data to export")
                return
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not file_path:
                return
            
            # Get headers
            headers = self.lte_tree['columns']
            
            # Create dataframe
            data = []
            for item in items:
                data.append(self.lte_tree.item(item, 'values'))
            
            df = pd.DataFrame(data, columns=headers)
            
            # Save to Excel
            df.to_excel(file_path, index=False)
            self.update_status(f"Exported {len(data)} LTE rows to {os.path.basename(file_path)}")
            messagebox.showinfo("Success", "LTE data exported successfully!")
            
        except Exception as e:
            logging.error(f"Error in export_lte_to_excel: {str(e)}")
            self.update_status(f"Export error: {str(e)}")
            messagebox.showerror("Export Error", f"Failed to export LTE data: {str(e)}")
    
    def export_5g_to_excel(self):
        """Export 5G parameters to Excel"""
        try:
            items = self.nr_tree.get_children()
            if not items:
                messagebox.showinfo("Info", "No 5G data to export")
                return
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not file_path:
                return
            
            # Get headers
            headers = self.nr_tree['columns']
            
            # Create dataframe
            data = []
            for item in items:
                data.append(self.nr_tree.item(item, 'values'))
            
            df = pd.DataFrame(data, columns=headers)
            
            # Save to Excel
            df.to_excel(file_path, index=False)
            self.update_status(f"Exported {len(data)} 5G rows to {os.path.basename(file_path)}")
            messagebox.showinfo("Success", "5G data exported successfully!")
            
        except Exception as e:
            logging.error(f"Error in export_5g_to_excel: {str(e)}")
            self.update_status(f"Export error: {str(e)}")
            messagebox.showerror("Export Error", f"Failed to export 5G data: {str(e)}")
    
    def export_results(self, items):
        """Export given items to Excel file"""
        try:
            if not items:
                messagebox.showinfo("Info", "No data to export")
                return
            
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not file_path:
                return
            
            # Get headers
            headers = self.tree['columns']
            
            # Create dataframe
            data = []
            for item in items:
                data.append(self.tree.item(item, 'values'))
            
            df = pd.DataFrame(data, columns=headers)
            
            # Save to Excel
            df.to_excel(file_path, index=False)
            self.update_status(f"Exported {len(data)} rows to {os.path.basename(file_path)}")
            messagebox.showinfo("Success", "Data exported successfully!")
            
        except Exception as e:
            logging.error(f"Error in export_results: {str(e)}")
            self.update_status(f"Export error: {str(e)}")
            messagebox.showerror("Export Error", f"Failed to export data: {str(e)}")
    
    # Distance calculator methods
    def add_point(self):
        """Add a named point to the list using separate lat/lon entries"""
        try:
            name = self.point_name_entry.get().strip()
            lat_str = self.lat_entry.get().strip()
            lon_str = self.lon_entry.get().strip()
            
            if not name:
                name = f"Point {len(self.points) + 1}"
            
            if not lat_str or not lon_str:
                messagebox.showwarning("Input Error", "Please enter both latitude and longitude")
                return
                
            lat = float(lat_str)
            lon = float(lon_str)
            
            # Add to list
            self.points.append((name, lat, lon))
            self.points_listbox.insert(tk.END, f"{name}: {lat:.6f}, {lon:.6f}")
            
            # Clear entries
            self.point_name_entry.delete(0, tk.END)
            self.lat_entry.delete(0, tk.END)
            self.lon_entry.delete(0, tk.END)
            
        except ValueError:
            messagebox.showerror("Input Error", "Invalid latitude or longitude format")
        except Exception as e:
            logging.error(f"Error in add_point: {str(e)}")
            messagebox.showerror("Error", f"Failed to add point: {str(e)}")
    
    def add_point_from_lat_long(self):
        """Add a named point to the list using a single lat,long entry"""
        try:
            lat_long_str = self.lat_long_entry.get().strip()
            name = self.lat_long_name_entry.get().strip()
            
            if not lat_long_str:
                messagebox.showwarning("Input Error", "Please enter latitude and longitude (e.g., 12.34,78.56)")
                return
            
            if not name:
                name = f"Point {len(self.points) + 1}"
            
            # Parse lat,long
            try:
                lat_str, lon_str = lat_long_str.split(',')
                lat = float(lat_str.strip())
                lon = float(lon_str.strip())
            except:
                raise ValueError("Invalid format")
            
            # Add to list
            self.points.append((name, lat, lon))
            self.points_listbox.insert(tk.END, f"{name}: {lat:.6f}, {lon:.6f}")
            
            # Clear entries
            self.lat_long_entry.delete(0, tk.END)
            self.lat_long_name_entry.delete(0, tk.END)
            
        except ValueError:
            messagebox.showerror("Input Error", "Invalid format. Please use 'latitude,longitude' (e.e., 12.34,78.56)")
        except Exception as e:
            logging.error(f"Error in add_point_from_lat_long: {str(e)}")
            messagebox.showerror("Error", f"Failed to add point: {str(e)}")
    
    def import_from_excel(self):
        """Import points from Excel file"""
        try:
            file_path = filedialog.askopenfilename(
                title="Select Excel File with Points",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            
            if not file_path:
                return
                
            # Read Excel file
            df = pd.read_excel(file_path)
            
            # Try to detect columns
            name_col = None
            lat_col = None
            lon_col = None
            
            for col in df.columns:
                col_lower = str(col).lower()
                if "name" in col_lower or "point" in col_lower:
                    name_col = col
                elif "lat" in col_lower:
                    lat_col = col
                elif "lon" in col_lower or "long" in col_lower:
                    lon_col = col
            
            # Validate columns
            if not lat_col or not lon_col:
                messagebox.showerror("Import Error", "Could not find latitude and longitude columns in the file")
                return
            
            # Add points
            count = 0
            for _, row in df.iterrows():
                name = row[name_col] if name_col and name_col in row and not pd.isna(row[name_col]) else f"Point {len(self.points) + 1}"
                lat = row[lat_col]
                lon = row[lon_col]
                
                if pd.isna(lat) or pd.isna(lon):
                    continue
                
                self.points.append((name, lat, lon))
                self.points_listbox.insert(tk.END, f"{name}: {lat:.6f}, {lon:.6f}")
                count += 1
            
            self.update_status(f"Imported {count} points from {os.path.basename(file_path)}")
            
        except Exception as e:
            logging.error(f"Error in import_from_excel: {str(e)}")
            messagebox.showerror("Import Error", f"Failed to import points: {str(e)}")
    
    def set_as_master(self):
        """Set the selected point as the master point (manual button press)"""
        try:
            selection = self.points_listbox.curselection()
            if not selection:
                messagebox.showinfo("Info", "Please select a point first")
                return
                
            index = selection[0]
            point = self.points[index]
            response = messagebox.askyesno("Set Master Point", f"Is '{point[0]}' your master point?")
            if response:
                self.master_point = point
                self.update_status(f"Set master point: {self.master_point[0]}")
                messagebox.showinfo("Master Point", f"Set '{self.master_point[0]}' as the master point")
        except Exception as e:
            logging.error(f"Error in set_as_master: {str(e)}")
            messagebox.showerror("Error", f"Failed to set master point: {str(e)}")
    
    def remove_point(self):
        """Remove selected point from list"""
        try:
            selection = self.points_listbox.curselection()
            if not selection:
                return
                
            index = selection[0]
            
            # If removing master point, clear it
            if self.master_point == self.points[index]:
                self.master_point = None
            
            self.points_listbox.delete(index)
            self.points.pop(index)
        except Exception as e:
            logging.error(f"Error in remove_point: {str(e)}")
            messagebox.showerror("Error", f"Failed to remove point: {str(e)}")
    
    def clear_points(self):
        """Clear all points"""
        try:
            self.points_listbox.delete(0, tk.END)
            self.points = []
            self.master_point = None
            self.distance_text.config(state=tk.NORMAL)
            self.distance_text.delete(1.0, tk.END)
            self.distance_text.config(state=tk.DISABLED)
        except Exception as e:
            logging.error(f"Error in clear_points: {str(e)}")
            messagebox.showerror("Error", f"Failed to clear points: {str(e)}")
    
    def move_point(self, direction):
        """Move a point up or down in the list"""
        try:
            selection = self.points_listbox.curselection()
            if not selection:
                return
                
            index = selection[0]
            if (direction < 0 and index == 0) or (direction > 0 and index == len(self.points) - 1):
                return
                
            new_index = index + direction
            # Get current point
            point = self.points[index]
            point_str = self.points_listbox.get(index)
            
            # Remove and reinsert
            self.points_listbox.delete(index)
            self.points.pop(index)
            
            self.points_listbox.insert(new_index, point_str)
            self.points.insert(new_index, point)
            
            # Select the moved item
            self.points_listbox.selection_set(new_index)
        except Exception as e:
            logging.error(f"Error in move_point: {str(e)}")
            messagebox.showerror("Error", f"Failed to move point: {str(e)}")
    
    def calculate_path_distances(self):
        """Calculate distances between all points in the path"""
        try:
            if len(self.points) < 2:
                messagebox.showinfo("Info", "At least two points are required for distance calculation")
                return
                
            total_distance = 0
            segment_distances = []
            
            # Calculate distances between consecutive points
            for i in range(len(self.points) - 1):
                name1, lat1, lon1 = self.points[i]
                name2, lat2, lon2 = self.points[i+1]
                
                distance_km = self.calculate_distance(lat1, lon1, lat2, lon2)
                total_distance += distance_km
                
                # Store segment result
                segment_distances.append(f"{name1} → {name2}: {distance_km:.3f} km")
            
            # Display results
            self.distance_text.config(state=tk.NORMAL)
            self.distance_text.delete(1.0, tk.END)
            
            # Add header
            self.distance_text.insert(tk.END, "Path Distances:\n", "header")
            
            # Add segment distances
            for segment in segment_distances:
                self.distance_text.insert(tk.END, segment + "\n")
            
            # Add total distance
            self.distance_text.insert(tk.END, f"\nTotal Distance: {total_distance:.3f} km\n", "bold")
            
            self.distance_text.config(state=tk.DISABLED)
            self.dist_notebook.select(0)  # Show results tab
        except Exception as e:
            logging.error(f"Error in calculate_path_distances: {str(e)}")
            messagebox.showerror("Error", f"Failed to calculate path distances: {str(e)}")
    
    def calculate_from_master(self):
        """Calculate distances from master point to all other points"""
        try:
            if not self.master_point:
                messagebox.showinfo("Info", "Please set a master point first by selecting a point and confirming 'Yes'")
                return
                
            if len(self.points) < 2:
                messagebox.showinfo("Info", "At least one additional point is required")
                return
                
            master_name, master_lat, master_lon = self.master_point
            total_distance = 0
            distances = []
            
            for name, lat, lon in self.points:
                if (name, lat, lon) == self.master_point:
                    continue
                    
                distance_km = self.calculate_distance(master_lat, master_lon, lat, lon)
                total_distance += distance_km
                distances.append(f"{master_name} to {name}: {distance_km:.3f} km")
            
            # Display results
            self.distance_text.config(state=tk.NORMAL)
            self.distance_text.delete(1.0, tk.END)
            
            # Add header
            self.distance_text.insert(tk.END, "Distances from Master Point:\n", "header")
            
            # Add distances
            for distance in distances:
                self.distance_text.insert(tk.END, distance + "\n")
            
            # Add total distance
            self.distance_text.insert(tk.END, f"\nTotal Distance: {total_distance:.3f} km\n", "bold")
            
            self.distance_text.config(state=tk.DISABLED)
            self.dist_notebook.select(0)  # Show results tab
        except Exception as e:
            logging.error(f"Error in calculate_from_master: {str(e)}")
            messagebox.showerror("Error", f"Failed to calculate from master: {str(e)}")
    
    def calculate_distance(self, lat1, lon1, lat2, lon2):
        """Calculate distance between two points using Haversine formula"""
        try:
            R = 6371.0  # Earth radius in kilometers
            
            lat1_rad = radians(lat1)
            lon1_rad = radians(lon1)
            lat2_rad = radians(lat2)
            lon2_rad = radians(lon2)
            
            dlon = lon2_rad - lon1_rad
            dlat = lat2_rad - lat1_rad
            
            a = sin(dlat / 2)**2 + cos(lat1_rad) * cos(lat2_rad) * sin(dlon / 2)**2
            c = 2 * atan2(sqrt(a), sqrt(1 - a))
            
            return R * c
        except Exception as e:
            logging.error(f"Error in calculate_distance: {str(e)}")
            return 0
    
    def open_google_maps(self):
        """Open Google Maps with the current points in default browser with names and distances"""
        try:
            if not self.points:
                messagebox.showinfo("Info", "No points to show on map")
                return
                
            # Create HTML content with all points and distances
            html_content = self.generate_map_html()
            
            # Save to temp file
            with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.html') as f:
                f.write(html_content)
                temp_path = f.name
            
            # Open in default browser
            webbrowser.open(f"file://{temp_path}")
        except Exception as e:
            logging.error(f"Error in open_google_maps: {str(e)}")
            messagebox.showerror("Error", f"Failed to open Google Maps: {str(e)}")
    
    def generate_map_html(self):
        """Generate HTML with Google Maps showing all points with names and distances"""
        try:
            # Create HTML content
            html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Network Points Map</title>
                <style>
                    body {{ margin: 0; padding: 0; }}
                    #map {{ height: 100vh; width: 100%; }}
                    #info-panel {{ position: absolute; top: 10px; left: 10px; background: white; padding: 10px; z-index: 1000; }}
                </style>
            </head>
            <body>
                <div id="info-panel"></div>
                <div id="map"></div>
                
                <script>
                    function initMap() {{
                        // Create map
                        const map = new google.maps.Map(document.getElementById("map"), {{
                            zoom: 12,
                            center: {{ lat: {self.points[0][1] if self.points else 0}, lng: {self.points[0][2] if self.points else 0} }},
                            mapTypeId: 'terrain'
                        }});
                        
                        // Create info window
                        const infoWindow = new google.maps.InfoWindow();
                        const infoPanel = document.getElementById("info-panel");
                        
                        // Array to store markers
                        const markers = [];
                        let infoContent = "<h3>Points and Distances</h3><ul>";
                        
                        // Add markers for each point
                        {self.generate_markers_js()}
                        
                        // Draw paths if needed
                        {self.generate_paths_js()}
                        
                        // Add master point info if exists
                        {self.generate_master_info_js()}
                        
                        infoContent += "</ul>";
                        infoPanel.innerHTML = infoContent;
                    }}
                </script>
                <script async defer
                    src="https://maps.googleapis.com/maps/api/js?key={self.api_key}&callback=initMap">
                </script>
            </body>
            </html>
            """
            return html
        except Exception as e:
            logging.error(f"Error in generate_map_html: {str(e)}")
            return "<html><body>Error generating map</body></html>"
    
    def generate_markers_js(self):
        """Generate JavaScript code for markers"""
        try:
            js_code = ""
            for i, (name, lat, lon) in enumerate(self.points):
                # Add marker
                js_code += f"""
                    // Marker for {name}
                    const marker{i} = new google.maps.Marker({{
                        position: {{ lat: {lat}, lng: {lon} }},
                        map: map,
                        title: "{name}",
                        label: "{i+1}",
                    }});
                    
                    // Info window for {name}
                    marker{i}.addListener('click', function() {{
                        infoWindow.setContent('<div><strong>{name}</strong><br>Lat: {lat:.6f}<br>Lon: {lon:.6f}</div>');
                        infoWindow.open(map, marker{i});
                    }});
                    
                    markers.push(marker{i});
                    infoContent += "<li><strong>{name}</strong> (Lat: {lat:.6f}, Lon: {lon:.6f})</li>";
                """
            return js_code
        except Exception as e:
            logging.error(f"Error in generate_markers_js: {str(e)}")
            return ""
    
    def generate_paths_js(self):
        """Generate JavaScript code for paths and distances"""
        try:
            if len(self.points) < 2:
                return ""
                
            js_code = "// Create path\n const pathCoordinates = ["
            
            # Add coordinates
            for name, lat, lon in self.points:
                js_code += f"{{lat: {lat}, lng: {lon}}},"
            
            js_code += "];\n"
            
            # Draw polyline
            js_code += """
            const path = new google.maps.Polyline({
                path: pathCoordinates,
                geodesic: true,
                strokeColor: "#FF0000",
                strokeOpacity: 1.0,
                strokeWeight: 2
            });
            path.setMap(map);
            """
            
            # Add distance markers
            js_code += "// Add distance markers\n"
            for i in range(len(self.points) - 1):
                name1, lat1, lon1 = self.points[i]
                name2, lat2, lon2 = self.points[i+1]
                distance = self.calculate_distance(lat1, lon1, lat2, lon2)
                
                # Midpoint for label
                mid_lat = (lat1 + lat2) / 2
                mid_lon = (lon1 + lon2) / 2
                
                js_code += f"""
                // Distance between {name1} and {name2}
                const distanceMarker{i} = new google.maps.Marker({{
                    position: {{ lat: {mid_lat}, lng: {mid_lon} }},
                    map: map,
                    icon: {{
                        path: google.maps.SymbolPath.CIRCLE,
                        scale: 0  // Hide marker, just show label
                    }},
                    label: {{
                        text: "{distance:.2f} km",
                        color: "#0000FF",
                        fontSize: "12px",
                        fontWeight: "bold"
                    }}
                }});
                """
            
            return js_code
        except Exception as e:
            logging.error(f"Error in generate_paths_js: {str(e)}")
            return ""
    
    def generate_master_info_js(self):
        """Generate JavaScript code for master point info"""
        try:
            if not self.master_point:
                return ""
                
            master_name, master_lat, master_lon = self.master_point
            js_code = f"""
            // Highlight master point
            const masterMarker = new google.maps.Marker({{
                position: {{ lat: {master_lat}, lng: {master_lon} }},
                map: map,
                title: "Master: {master_name}",
                label: "M",
                icon: {{
                    url: "http://maps.google.com/mapfiles/ms/icons/blue-dot.png"
                }}
            }});
            
            // Add master distances
            infoContent += "<h4>Distances from Master Point:</h4><ul>";
            """
            
            for name, lat, lon in self.points:
                if (name, lat, lon) == self.master_point:
                    continue
                    
                distance = self.calculate_distance(master_lat, master_lon, lat, lon)
                js_code += f"""
                infoContent += "<li>{master_name} to {name}: {distance:.3f} km</li>";
                """
            
            js_code += "infoContent += \"</ul>\";"
            return js_code
        except Exception as e:
            logging.error(f"Error in generate_master_info_js: {str(e)}")
            return ""
    
    def show_map(self):
        """Show map with all points"""
        try:
            if not self.points:
                messagebox.showinfo("Info", "No points to show on map")
                return
                
            # Check API key
            if not self.api_key:
                messagebox.showwarning("API Key Missing", 
                    "Google Maps API key is required for map visualization. Please set the environment variable GOOGLE_MAPS_API_KEY")
                return
            
            # Create Google Maps URL
            base_url = "https://maps.googleapis.com/maps/api/staticmap?"
            size = "800x600"
            maptype = self.map_type.get()
            zoom = self.map_zoom.get()
            scale = "2"  # Higher resolution
            
            # Add markers
            markers = []
            for i, (name, lat, lon) in enumerate(self.points):
                color = "blue" if self.master_point and (name, lat, lon) == self.master_point else "red"
                label = chr(65 + i)  # A, B, C, ...
                markers.append(f"markers=color:{color}%7Clabel:{label}%7C{lat},{lon}")
            
            # Add path if there are multiple points
            path = ""
            if len(self.points) > 1:
                path_points = "|".join(f"{lat},{lon}" for _, lat, lon in self.points)
                path = f"&path=color:0xff0000%7Cweight:5%7C{path_points}"
            
            # Create URL
            url = f"{base_url}size={size}&maptype={maptype}&zoom={zoom}&scale={scale}&{('&'.join(markers))}{path}&key={self.api_key}"
            
            # Try to display map image
            response = requests.get(url)
            if response.status_code != 200:
                error_details = response.text
                if len(error_details) > 200:
                    error_details = error_details[:200] + "..."
                raise Exception(f"Google Maps API returned {response.status_code}: {error_details}")
            
            img_data = response.content
            img = Image.open(io.BytesIO(img_data))
            photo = ImageTk.PhotoImage(img)
            
            self.map_label.config(image=photo)
            self.map_label.image = photo  # Keep reference
            
            # Create legend
            legend_text = "Map Legend:\n"
            for i, (name, _, _) in enumerate(self.points):
                label = chr(65 + i)
                legend_text += f"{label}: {name}\n"
            
            if self.master_point:
                legend_text += f"\nMaster Point: {self.master_point[0]} (blue)"
            
            # Add legend to map
            self.map_label.config(text=legend_text, compound=tk.BOTTOM)
            self.dist_notebook.select(1)  # Show map tab
            
        except Exception as e:
            logging.error(f"Error in show_map: {str(e)}")
            messagebox.showerror("Map Error", f"Failed to load map: {str(e)}")
            # Fallback to opening in browser
            self.open_google_maps()
    
    def use_for_distance(self):
        """Use selected point from results for distance calculation"""
        try:
            selection = self.tree.selection()
            if not selection:
                return
                
            item = selection[0]
            values = self.tree.item(item, 'values')
            col_names = self.tree['columns']
            
            # Find latitude and longitude columns
            lat_idx = col_names.index("LATITUDE")
            lon_idx = col_names.index("LONGITUDE")
            
            lat_str = values[lat_idx]
            lon_str = values[lon_idx]
            
            # Get name from the record
            name_idx = col_names.index("Site") if "Site" in col_names else None
            name = values[name_idx] if name_idx is not None else f"Point {len(self.points) + 1}"
            
            # Clean and convert
            lat = float(lat_str) if lat_str else None
            lon = float(lon_str) if lon_str else None
            
            if lat is None or lon is None:
                raise ValueError("Missing coordinates")
                
            # Add to points
            self.points.append((name, lat, lon))
            self.points_listbox.insert(tk.END, f"{name}: {lat:.6f}, {lon:.6f}")
            
        except (ValueError, IndexError) as e:
            messagebox.showerror("Error", "Could not get valid coordinates from this row")
        except Exception as e:
            logging.error(f"Error in use_for_distance: {str(e)}")
            messagebox.showerror("Error", f"Failed to use for distance: {str(e)}")
    
    def prompt_master_point(self, event):
        """Prompt user to confirm if selected point should be master point on left-click"""
        try:
            selection = self.points_listbox.curselection()
            if selection:
                index = selection[0]
                point = self.points[index]
                response = messagebox.askyesno("Set Master Point", f"Is '{point[0]}' your master point?")
                if response:
                    self.master_point = point
                    self.update_status(f"Set master point: {self.master_point[0]}")
                    messagebox.showinfo("Master Point", f"Set '{self.master_point[0]}' as the master point")
        except Exception as e:
            logging.error(f"Error in prompt_master_point: {str(e)}")
            messagebox.showerror("Error", f"Failed to set master point: {str(e)}")
    
    def handle_nr_right_click(self, event):
        """Handle right-click in 5G treeview"""
        try:
            region = self.nr_tree.identify("region", event.x, event.y)
            if region == "heading":
                col_id = self.nr_tree.identify_column(event.x)
                col_index = int(col_id.replace('#', '')) - 1
                columns = self.nr_tree['columns']
                if col_index < len(columns):
                    self.current_col_index = col_index
                    self.current_col_name = columns[col_index]
                    
                    # Create context menu for column
                    menu = tk.Menu(self.root, tearoff=0)
                    menu.add_command(label="Delete Column", command=self.delete_5g_column)
                    menu.add_command(label="Rename Column", command=self.rename_5g_column)
                    menu.tk_popup(event.x_root, event.y_root)
            else:
                # For rows/cells
                self.current_col = self.nr_tree.identify_column(event.x)
                self.current_row = self.nr_tree.identify_row(event.y)
                selection = self.nr_tree.selection()
                if selection:
                    self.nr_context_menu.entryconfig("Copy with Headers", state=tk.NORMAL)
                    self.nr_context_menu.entryconfig("Export Selected to Excel", state=tk.NORMAL)
                else:
                    self.nr_context_menu.entryconfig("Copy with Headers", state=tk.DISABLED)
                    self.nr_context_menu.entryconfig("Export Selected to Excel", state=tk.DISABLED)
                self.nr_context_menu.tk_popup(event.x_root, event.y_root)
        except Exception as e:
            logging.error(f"Error in handle_nr_right_click: {str(e)}")
    
    def rename_5g_column(self):
        """Rename selected column in 5G treeview"""
        try:
            if hasattr(self, 'current_col_name'):
                new_name = simpledialog.askstring("Rename Column", "New column name:", initialvalue=self.current_col_name)
                if new_name and new_name != self.current_col_name:
                    current_columns = list(self.nr_tree['columns'])
                    if new_name in current_columns:
                        messagebox.showwarning("Duplicate", "Column name already exists")
                        return
                    idx = current_columns.index(self.current_col_name)
                    current_columns[idx] = new_name
                    self.rebuild_5g_tree(current_columns)
        except Exception as e:
            logging.error(f"Error in rename_5g_column: {str(e)}")
            messagebox.showerror("Error", f"Failed to rename column: {str(e)}")
    
    def copy_nr_cell(self):
        """Copy single cell value from 5G treeview"""
        try:
            if not hasattr(self, 'current_row') or not self.current_row:
                return
            item = self.current_row
            col_idx = int(self.current_col.replace('#', '')) - 1
            values = self.nr_tree.item(item, 'values')
            if col_idx < len(values):
                cell_value = str(values[col_idx])
                self.root.clipboard_clear()
                self.root.clipboard_append(cell_value)
                self.update_status("Copied cell value to clipboard")
        except Exception as e:
            logging.error(f"Error in copy_nr_cell: {str(e)}")
            messagebox.showerror("Error", f"Failed to copy cell: {str(e)}")
    
    def copy_5g_with_headers(self):
        """Copy selected rows with headers from 5G treeview"""
        try:
            selected = self.nr_tree.selection()
            if not selected:
                return
            headers = self.nr_tree['columns']
            header_text = "\t".join(headers)
            data_rows = []
            for item in selected:
                values = self.nr_tree.item(item, 'values')
                data_rows.append("\t".join(map(str, values)))
            clipboard_text = header_text + "\n" + "\n".join(data_rows)
            self.root.clipboard_clear()
            self.root.clipboard_append(clipboard_text)
            self.update_status("Copied selected rows with headers to clipboard")
        except Exception as e:
            logging.error(f"Error in copy_5g_with_headers: {str(e)}")
            messagebox.showerror("Error", f"Failed to copy: {str(e)}")
    
    def export_5g_selected_to_excel(self):
        """Export selected rows from 5G treeview to Excel"""
        try:
            selected = self.nr_tree.selection()
            if not selected:
                messagebox.showinfo("Info", "No rows selected for export")
                return
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if not file_path:
                return
            headers = self.nr_tree['columns']
            data = []
            for item in selected:
                data.append(self.nr_tree.item(item, 'values'))
            df = pd.DataFrame(data, columns=headers)
            df.to_excel(file_path, index=False)
            self.update_status(f"Exported {len(data)} rows to {os.path.basename(file_path)}")
            messagebox.showinfo("Success", "Data exported successfully!")
        except Exception as e:
            logging.error(f"Error in export_5g_selected_to_excel: {str(e)}")
            messagebox.showerror("Error", f"Export failed: {str(e)}")
    
    def delete_5g_column(self):
        """Delete selected column from 5G treeview"""
        try:
            if hasattr(self, 'current_col_name') and self.current_col_name:
                # Get current columns
                current_columns = list(self.nr_tree['columns'])
                
                # Skip required columns
                required_columns = {"Source", "USID", "SITE", "NRCELL_NAME"}
                if self.current_col_name in required_columns:
                    messagebox.showwarning("Cannot Delete", f"Cannot delete required column: {self.current_col_name}")
                    return
                
                # Remove the selected column
                if self.current_col_name in current_columns:
                    current_columns.remove(self.current_col_name)
                    
                    # Rebuild tree with updated columns
                    self.rebuild_5g_tree(current_columns)
        except Exception as e:
            logging.error(f"Error in delete_5g_column: {str(e)}")
            messagebox.showerror("Error", f"Failed to delete column: {str(e)}")
    
    def add_5g_column(self):
        """Add a new column to the 5G treeview"""
        try:
            # Get available data sources
            data_sources = ["LTE", "5GNR", "5GNR_BBU"]
            
            # Create dialog for column selection
            dialog = tk.Toplevel(self.root)
            dialog.title("Add Column")
            dialog.geometry("400x300")
            dialog.transient(self.root)
            dialog.grab_set()
            
            # Source selection
            ttk.Label(dialog, text="Select Data Source:").pack(pady=5)
            source_var = tk.StringVar(value=data_sources[0])
            source_combo = ttk.Combobox(dialog, textvariable=source_var, values=data_sources, state="readonly")
            source_combo.pack(side=tk.LEFT, padx=5)
            
            # Column selection
            ttk.Label(dialog, text="Select Column:").pack(pady=5)
            column_listbox = tk.Listbox(dialog)
            column_listbox.pack(fill=tk.BOTH, expand=True, pady=5)
            
            # Populate initial columns
            self.populate_column_list(column_listbox, data_sources[0])
            
            # Update column list when source changes
            def on_source_change(event):
                self.populate_column_list(column_listbox, source_var.get())
            
            source_combo.bind("<<ComboboxSelected>>", on_source_change)
            
            # Add button
            def add_selected_column():
                selected = column_listbox.curselection()
                if selected:
                    column_name = column_listbox.get(selected[0])
                    current_columns = list(self.nr_tree['columns'])
                    
                    # Skip if column already exists
                    if column_name in current_columns:
                        messagebox.showinfo("Info", f"Column '{column_name}' already exists")
                        return
                    
                    # Add new column
                    current_columns.append(column_name)
                    self.rebuild_5g_tree(current_columns)
                    dialog.destroy()
            
            ttk.Button(dialog, text="Add", command=add_selected_column).pack(pady=10)
            
        except Exception as e:
            logging.error(f"Error in add_5g_column: {str(e)}")
            messagebox.showerror("Error", f"Failed to add column: {str(e)}")
    
    def rebuild_5g_tree(self, new_columns):
        """Rebuild the 5G treeview with new columns"""
        try:
            # Create a new tree with updated columns
            old_tree = self.nr_tree
            tree_container = old_tree.master  # This is the container frame
            results_frame = tree_container.master  # Parent of tree_container
            
            # Create new container
            new_container = ttk.Frame(results_frame)
            new_container.pack(fill=tk.BOTH, expand=True)
            
            # Create new tree
            self.nr_tree = ttk.Treeview(new_container, columns=new_columns, show="headings", selectmode="extended")
            vsb = ttk.Scrollbar(new_container, orient="vertical", command=self.nr_tree.yview)
            hsb = ttk.Scrollbar(new_container, orient="horizontal", command=self.nr_tree.xview)
            self.nr_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            
            # Configure columns
            col_widths = {
                "Source": 50, "USID": 70, "SITE": 90, "NRCELL_NAME": 110, 
                "Digital Tilt": 70, "Power": 50, "PCI": 40,
                "ADMINISTRATIVESTATE": 90, "CELLBARRED": 70, 
                "CELLRESERVEDFOROPERATOR": 90, "OPERATIONALSTATE": 90,
                "CELLRANGE": 70, "SSBFREQUENCY": 70, "CONFIGURATION": 70
            }
            
            for col in new_columns:
                self.nr_tree.heading(col, text=col)
                self.nr_tree.column(col, width=col_widths.get(col, 70), anchor=tk.CENTER)
            
            # Layout new tree and scrollbars
            self.nr_tree.grid(row=0, column=0, sticky="nsew")
            vsb.grid(row=0, column=1, sticky="ns")
            hsb.grid(row=1, column=0, sticky="ew")
            
            # Configure grid weights for container
            new_container.grid_rowconfigure(0, weight=1)
            new_container.grid_columnconfigure(0, weight=1)
            
            # Rebind right-click
            self.nr_tree.bind("<Button-3>", self.handle_nr_right_click)
            
            # Repopulate with existing records
            if hasattr(self, 'matched_records') and self.matched_records:
                for tech, record in self.matched_records:
                    if tech == "5GNR":
                        self.add_to_5g_tree(tech, record)
            
            # Destroy old container and tree
            tree_container.destroy()
            
            self.update_status(f"5G Parameters tab updated with {len(new_columns)} columns")
            
        except Exception as e:
            logging.error(f"Error in rebuild_5g_tree: {str(e)}")
            messagebox.showerror("Error", f"Failed to rebuild tree: {str(e)}")

def main():
    """Main function with enhanced exception handling"""
    try:
        root = tk.Tk()
        app = NetworkSearchApp(root)
        root.mainloop()
    except Exception as e:
        error_msg = f"Critical Error: {str(e)}\n\n{traceback.format_exc()}"
        logging.critical(error_msg)
        messagebox.showerror("Fatal Error", f"Application crashed: {str(e)}")
        # Attempt to log the error to a file
        try:
            with open("network_search_crash.log", "a") as f:
                f.write(f"{datetime.now()}: {error_msg}\n")
        except:
            pass

if __name__ == "__main__":
    main()