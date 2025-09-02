import pandas as pd
import os
import streamlit as st
import logging
from datetime import datetime
import openpyxl
import math
from math import radians, sin, cos, sqrt, atan2
import webbrowser
import io
import requests
import tempfile
import base64

# Configure logging
logging.basicConfig(filename='network_search.log', level=logging.ERROR,
                    format='%(asctime)s - %(levelname)s - %(message)s')

class NetworkSearchApp:
    def __init__(self):
        # Initialize data structures in session state
        if 'lte_data' not in st.session_state:
            st.session_state.lte_data = pd.DataFrame()
        if 'nr_data' not in st.session_state:
            st.session_state.nr_data = pd.DataFrame()
        if 'bbu_data' not in st.session_state:
            st.session_state.bbu_data = pd.DataFrame()
        if 'file_paths' not in st.session_state:
            st.session_state.file_paths = []
        if 'matched_records' not in st.session_state:
            st.session_state.matched_records = []
        if 'points' not in st.session_state:
            st.session_state.points = []
        if 'master_point' not in st.session_state:
            st.session_state.master_point = None
        if 'vdt_data' not in st.session_state:
            st.session_state.vdt_data = pd.DataFrame(columns=["LTE Site", "NR Site"])
        if 'auto_generate' not in st.session_state:
            st.session_state.auto_generate = True
        if 'project_name' not in st.session_state:
            st.session_state.project_name = "ATT_STX_253"
        if 'search_type' not in st.session_state:
            st.session_state.search_type = "USID"
        if 'search_value' not in st.session_state:
            st.session_state.search_value = ""
        if 'lte_cr_type' not in st.session_state:
            st.session_state.lte_cr_type = "cellRange"
        if 'nr_cr_type' not in st.session_state:
            st.session_state.nr_cr_type = "digitalTilt"
        if 'map_type' not in st.session_state:
            st.session_state.map_type = "terrain"
        if 'map_zoom' not in st.session_state:
            st.session_state.map_zoom = 12
        if 'distance_results' not in st.session_state:
            st.session_state.distance_results = ""
        if 'lte_columns' not in st.session_state:
            st.session_state.lte_columns = [
                "Source", "Site", "cell", "CELLRANGE", "CRSGAIN", "QRXLEVMIN", "EARFCNDL"
            ]
        if 'nr_columns' not in st.session_state:
            st.session_state.nr_columns = [
                "Source", "USID", "SITE", "NRCELL_NAME", "Digital Tilt", "Power", "PCI",
                "ADMINISTRATIVESTATE", "CELLBARRED", "CELLRESERVEDFOROPERATOR",
                "OPERATIONALSTATE", "CELLRANGE", "SSBFREQUENCY", "CONFIGURATION"
            ]

        # Market mapping
        self.market_mapping = {
            "ATT_ARK1": "ATT_ARK_253",
            "ATT_NoCAL1": "ATT_NoCAL_253",
            "ATT_SoCAL1": "ATT_SoCAL_253",
            "ATT_STX": "ATT_STX_253"
        }

        # Enhanced mappings
        self.mappings = {
            "LTE": {
                "USID": ["REMOTE_USID", "CSS_USID", "USID"],
                "ENBID": ["ENBID"],
                "cell ID": ["CELLID"],
                "Site": ["MECONTEXT_ID", "SITE", "OSS_ENodeB"],
                "Azumuth": ["ATOLL_AZIMUTH", "AZIMUTH", "Atoll_AZIMUT", "Atoll_Az"],
                "Digital Tilt": ["DIGITALTILT", "DIGITAL_TILT", "TILT"],
                "cell": ["EUTRAN_CELL_FDD_ID", "CELL", "OSS_EUTRAN_CELL_FDD_ID", "CELL_NAME"],
                "height(Meter)": ["Atoll_HEIGHT_m", "HEIGHT", "ANTENNA_HEIGHT", "HEIGHT_M"],
                "PCI": ["PHYSICALLAYERCELLID", "PCI", "OSS_PCI", "Atoll_PCI"],
                "Power": ["CONFIGUREDMAXTXPOWER", "TX_POWER", "OSS_CONFIGUREDMAXTXPOWER"],
                "LATITUDE": ["LATITUDE", "LAT"],
                "LONGITUDE": ["LONGITUDE", "LON", "LONG"],
                "ADMINISTRATIVESTATE": ["ADMINISTRATIVE_STATE", "ADMIN_STATE"],
                "OPERATIONALSTATE": ["OPERATIONALSTATE", "OP_STATE"],
                "CELLRANGE": ["CELLRANGE"],
                "CRSGAIN": ["CRSGAIN"],
                "QRXLEVMIN": ["QRXLEVMIN"],
                "EARFCNDL": ["EARFCNDL"],
                "Electrical Tilt": ["ELECTRICAL_TILT", "Atoll_ET", "E_TILT"],
                "ED_Market": ["ED_MARKET", "ED_Market", "EDMARKET"]
            },
            "5GNR": {
                "USID": ["CSS_USID", "REMOTE_USID", "USID", "BBU_USID"],
                "NIC": ["NCI"],
                "gnb ID": ["GNBID", "GNB_ID"],
                "Site": ["CTS_COMMON_ID", "GNB_NAME", "GNODEB"],
                "Azumuth": ["Atoll_AZIMUT", "AZIMUTH", "ATOLL_AZIMUTH"],
                "Digital Tilt": ["DIGITALTILT", "DIGITAL_TILT", "USEDDIGITALTILT"],
                "cell": ["NRCELLDUID", "CELL", "NRCELL_NAME"],
                "height(Meter)": ["Atoll_HEIGHT_m", "HEIGHT", "ANTENNA_HEIGHT", "ATOLL_ANTENNA_HEIGHT"],
                "PCI": ["NRPCI", "PCI"],
                "Power": ["CONFIGUREDMAXTXPOWER", "TX_POWER", "POWER"],
                "LATITUDE": ["LAT", "LATITUDE"],
                "LONGITUDE": ["LONG", "LONGITUDE"],
                "ADMINISTRATIVESTATE": ["ADMINISTRATIVESTATE", "ADMIN_STATE"],
                "OPERATIONALSTATE": ["OPERATIONALSTATE", "OP_STATE"],
                "CELLBARRED": ["CELLBARRED"],
                "CELLRESERVEDFOROPERATOR": ["CELLRESERVEDFOROPERATOR"],
                "CELLRANGE": ["CELLRANGE"],
                "SSBFREQUENCY": ["SSBFREQUENCY"],
                "ARFCNDL": ["ARFCNDL", "ARFCN_DL", "DL_ARFCN", "NR_ARFCNDL"],
                "CONFIGURATION": ["CONFIGURATION"],
                "Electrical Tilt": ["ELECTRICAL_TILT", "Atoll_ET", "E_TILT"],
                "ED_Market": ["ED_MARKET", "ED_Market", "EDMARKET"],
                "BBU_TECH": ["BBU_TECH"],
                "GNB_SA_STATE": ["GNB_SA_STATE"],
                "CELL_SA_STATE": ["CELL_SA_STATE"],
                "CELL_TYPE": ["CELL_TYPE"],
                "ON_AIR": ["ON_AIR"],
                "DSS_LTECELL": ["DSS_LTECELL"],
                "NRTAC": ["NRTAC"]
            },
            "5GNR_BBU": {
                "USID": ["USID", "REMOTE_USID", "CSS_USID"],
                "NRCELL_NAME": ["NRCELLDUID", "CELL", "NRCELL_NAME"],
                "CONFIGURATION": ["CONFIGURATION"],
                "ED_Market": ["ED_MARKET", "ED_Market", "EDMARKET"]
            }
        }

        # Google Maps API Key
        self.api_key = os.environ.get("GOOGLE_MAPS_API_KEY", "")

        # Create UI
        self.create_widgets()

    def update_status(self, message):
        """Update status with timestamp"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        st.session_state.status = f"[{timestamp}] {message}"
        st.write(st.session_state.status)

    def create_widgets(self):
        """Create Streamlit UI"""
        st.title("Network Data Search Tool")

        # Data Loading Section
        st.header("Data Loading")
        uploaded_files = st.file_uploader("Upload Multiple Files", accept_multiple_files=True, type=["xlsx", "xls"])
        if uploaded_files:
            st.session_state.file_paths = [file.name for file in uploaded_files]
            st.write("Uploaded Files:", ", ".join(st.session_state.file_paths))
        if st.button("Load Selected Data"):
            self.load_data(uploaded_files)

        # Search Section
        st.header("Search")
        col1, col2, col3 = st.columns([1, 2, 1])
        with col1:
            st.session_state.search_type = st.selectbox("Search By:", ["USID", "NIC", "gnb ID", "ENBID", "cell ID", "Site"], key="search_type")
        with col2:
            st.session_state.search_value = st.text_input("Value:", key="search_value")
        with col3:
            if st.button("Search"):
                self.perform_search()

        # Tabs
        tabs = st.tabs(["Main Results", "LTE Parameters", "5G Parameters", "VDT Sheet", "Distance Calculator"])

        # Main Results Tab
        with tabs[0]:
            self.create_main_tab()

        # LTE Parameters Tab
        with tabs[1]:
            self.create_lte_tab()

        # 5G Parameters Tab
        with tabs[2]:
            self.create_5g_tab()

        # VDT Sheet Tab
        with tabs[3]:
            self.create_vdt_tab()

        # Distance Calculator Tab
        with tabs[4]:
            self.create_distance_tab()

    def create_main_tab(self):
        """Create main results tab"""
        st.subheader("Main Results")
        columns = [
            "Source", "NIC", "gnb ID", "ENBID", "cell ID", "USID", "Site",
            "Azumuth", "Digital Tilt", "cell", "height(Meter)", "PCI", "Power",
            "LATITUDE", "LONGITUDE", "ADMINISTRATIVESTATE", "OPERATIONALSTATE"
        ]
        if st.session_state.matched_records:
            main_data = []
            for tech, record in st.session_state.matched_records:
                row = {"Source": tech}
                for key in self.mappings[tech]:
                    row[key] = self.get_column_value(record, self.mappings[tech][key], tech)
                main_data.append(row)
            df = pd.DataFrame(main_data, columns=columns)
            st.dataframe(df, use_container_width=True)
            if st.button("Export to Excel", key="main_export"):
                self.export_to_excel(df, "Main_Results.xlsx")
            if st.button("Clear Results", key="main_clear"):
                self.clear_results()

    def create_lte_tab(self):
        """Create LTE parameters tab"""
        st.subheader("LTE Parameters")
        col1, col2, col3, col4 = st.columns([1, 1, 1, 1])
        with col1:
            if st.button("Add Column", key="lte_add_col"):
                self.add_lte_column()
        with col2:
            st.session_state.lte_cr_type = st.selectbox("Generate CR for:", ["cellRange", "crsGain", "Electrical Tilt"], key="lte_cr_type")
        with col3:
            if st.button("Generate CR", key="lte_cr"):
                self.generate_lte_cr()
        with col4:
            if st.button("Export LTE Data to Excel", key="lte_export"):
                self.export_lte_to_excel()
        if st.session_state.matched_records:
            lte_data = []
            for tech, record in st.session_state.matched_records:
                if tech == "LTE":
                    row = {
                        "Source": "LTE",
                        "Site": self.get_column_value(record, self.mappings["LTE"]["Site"], "LTE"),
                        "cell": self.get_column_value(record, self.mappings["LTE"]["cell"], "LTE"),
                        "CELLRANGE": self.get_column_value(record, self.mappings["LTE"]["CELLRANGE"], "LTE"),
                        "CRSGAIN": self.get_column_value(record, self.mappings["LTE"]["CRSGAIN"], "LTE"),
                        "QRXLEVMIN": self.get_column_value(record, self.mappings["LTE"]["QRXLEVMIN"], "LTE"),
                        "EARFCNDL": self.get_column_value(record, self.mappings["LTE"]["EARFCNDL"], "LTE")
                    }
                    for col in st.session_state.lte_columns:
                        if col not in row:
                            row[col] = self.get_column_value(record, [col], "LTE")
                    lte_data.append(row)
            if lte_data:
                df = pd.DataFrame(lte_data, columns=st.session_state.lte_columns)
                st.dataframe(df, use_container_width=True)
                if st.button("Export Selected to Excel", key="lte_export_selected"):
                    self.export_lte_selected_to_excel(df)

    def create_5g_tab(self):
        """Create 5G parameters tab"""
        st.subheader("5G Parameters")
        col1, col2, col3, col4 = st.columns([1, 1, 1, 1])
        with col1:
            if st.button("Add Column", key="nr_add_col"):
                self.add_5g_column()
        with col2:
            st.session_state.nr_cr_type = st.selectbox("Generate CR for:", ["digitalTilt", "cellRange"], key="nr_cr_type")
        with col3:
            if st.button("Generate 5G CR", key="nr_cr"):
                self.generate_5g_cr()
        with col4:
            if st.button("Export 5G Data to Excel", key="nr_export"):
                self.export_5g_to_excel()
        if st.session_state.matched_records:
            nr_data = []
            for tech, record in st.session_state.matched_records:
                if tech == "5GNR":
                    row = {
                        "Source": "5GNR",
                        "USID": self.get_column_value(record, self.mappings["5GNR"]["USID"], "5GNR"),
                        "SITE": self.get_column_value(record, self.mappings["5GNR"]["Site"], "5GNR"),
                        "NRCELL_NAME": self.get_column_value(record, self.mappings["5GNR"]["cell"], "5GNR"),
                        "Digital Tilt": self.get_column_value(record, self.mappings["5GNR"]["Digital Tilt"], "5GNR"),
                        "Power": self.get_column_value(record, self.mappings["5GNR"]["Power"], "5GNR"),
                        "PCI": self.get_column_value(record, self.mappings["5GNR"]["PCI"], "5GNR"),
                        "ADMINISTRATIVESTATE": self.get_column_value(record, self.mappings["5GNR"]["ADMINISTRATIVESTATE"], "5GNR"),
                        "CELLBARRED": self.get_column_value(record, self.mappings["5GNR"]["CELLBARRED"], "5GNR"),
                        "CELLRESERVEDFOROPERATOR": self.get_column_value(record, self.mappings["5GNR"]["CELLRESERVEDFOROPERATOR"], "5GNR"),
                        "OPERATIONALSTATE": self.get_column_value(record, self.mappings["5GNR"]["OPERATIONALSTATE"], "5GNR"),
                        "CELLRANGE": self.get_column_value(record, self.mappings["5GNR"]["CELLRANGE"], "5GNR"),
                        "SSBFREQUENCY": self.get_column_value(record, self.mappings["5GNR"]["SSBFREQUENCY"], "5GNR"),
                        "CONFIGURATION": self.get_column_value(record, self.mappings["5GNR"]["CONFIGURATION"], "5GNR")
                    }
                    for col in st.session_state.nr_columns:
                        if col not in row:
                            row[col] = self.get_column_value(record, [col], "5GNR")
                    nr_data.append(row)
            if nr_data:
                df = pd.DataFrame(nr_data, columns=st.session_state.nr_columns)
                st.dataframe(df, use_container_width=True)
                if st.button("Export Selected to Excel", key="nr_export_selected"):
                    self.export_5g_selected_to_excel(df)

    def create_vdt_tab(self):
        """Create VDT sheet tab"""
        st.subheader("VDT Sheet")
        col1, col2, col3 = st.columns([1, 1, 1])
        with col1:
            st.session_state.project_name = st.selectbox("Project Name:", list(self.market_mapping.values()), key="project_name")
        with col2:
            if st.button("Generate VDT Report", key="vdt_report"):
                self.generate_vdt_report()
        with col3:
            st.session_state.auto_generate = st.checkbox("Auto-Generate on Search", value=st.session_state.auto_generate, key="auto_generate")
        if not st.session_state.vdt_data.empty:
            st.dataframe(st.session_state.vdt_data, use_container_width=True)

    def create_distance_tab(self):
        """Create distance calculator tab"""
        st.subheader("Distance Calculator")
        col1, col2 = st.columns([1, 2])
        with col1:
            st.write("Add Point")
            point_name = st.text_input("Point Name:", key="point_name")
            lat = st.number_input("Latitude:", format="%.6f", key="lat")
            lon = st.number_input("Longitude:", format="%.6f", key="lon")
            if st.button("Add Point", key="add_point"):
                self.add_point(point_name, lat, lon)
            st.write("Add from Lat,Long")
            lat_long = st.text_input("Lat,Long (e.g., 12.34,78.56):", key="lat_long")
            lat_long_name = st.text_input("Point Name for Lat,Long:", key="lat_long_name")
            if st.button("Add Lat,Long", key="add_lat_long"):
                self.add_lat_long(lat_long, lat_long_name)
            uploaded_points = st.file_uploader("Import Points from Excel", type=["xlsx", "xls"], key="points_upload")
            if uploaded_points:
                self.import_points_from_excel(uploaded_points)
            if st.session_state.points:
                st.write("Points List:")
                points_df = pd.DataFrame(st.session_state.points, columns=["Name", "Latitude", "Longitude"])
                selected = st.multiselect("Select Points:", points_df["Name"].tolist(), key="points_select")
                if st.button("Set as Master", key="set_master"):
                    if selected:
                        point = next(p for p in st.session_state.points if p[0] == selected[0])
                        self.set_master_point(point)
                if st.button("Remove Selected", key="remove_points"):
                    self.remove_points(selected)
                if st.button("Clear All Points", key="clear_points"):
                    self.clear_points()
        with col2:
            dist_tabs = st.tabs(["Results", "Map Visualization"])
            with dist_tabs[0]:
                if st.button("Calculate Path Distances", key="calc_path"):
                    self.calculate_path_distances()
                if st.button("Calculate from Master", key="calc_master"):
                    self.calculate_from_master()
                if st.session_state.distance_results:
                    st.text_area("Distance Results:", st.session_state.distance_results, height=300, key="distance_results")
            with dist_tabs[1]:
                col_map, col_zoom = st.columns([1, 1])
                with col_map:
                    st.session_state.map_type = st.selectbox("Map Type:", ["roadmap", "satellite", "hybrid", "terrain"], key="map_type")
                with col_zoom:
                    st.session_state.map_zoom = st.number_input("Zoom:", min_value=1, max_value=20, value=st.session_state.map_zoom, key="map_zoom")
                if st.button("Show Map", key="show_map"):
                    self.show_map()
                if st.session_state.points:
                    st.write("Map Legend:")
                    for i, (name, _, _) in enumerate(st.session_state.points):
                        st.write(f"{chr(65 + i)}: {name}")
                    if st.session_state.master_point:
                        st.write(f"Master Point: {st.session_state.master_point[0]} (blue)")

    def load_data(self, files):
        """Load data from uploaded files"""
        try:
            new_lte_data = []
            new_nr_data = []
            new_bbu_data = []
            for file in files:
                df = pd.read_excel(file)
                file_name = file.name.lower()
                if "lte" in file_name:
                    new_lte_data.append(df)
                elif "nr" in file_name:
                    new_nr_data.append(df)
                elif "bbu" in file_name:
                    new_bbu_data.append(df)
            st.session_state.lte_data = pd.concat(new_lte_data, ignore_index=True) if new_lte_data else pd.DataFrame()
            st.session_state.nr_data = pd.concat(new_nr_data, ignore_index=True) if new_nr_data else pd.DataFrame()
            st.session_state.bbu_data = pd.concat(new_bbu_data, ignore_index=True) if new_bbu_data else pd.DataFrame()
            self.update_status(f"Loaded {len(files)} files")
            st.success("Data loading completed!")
        except Exception as e:
            logging.error(f"Error in load_data: {str(e)}")
            self.update_status(f"Error loading files: {str(e)}")
            st.error(f"Error loading files: {str(e)}")

    def get_column_value(self, record, possible_names, tech):
        """Get column value from record using possible names"""
        for name in possible_names:
            if name in record and pd.notna(record[name]):
                return str(record[name])
        if tech == "5GNR" and possible_names[0] == "CONFIGURATION":
            usid = self.get_column_value(record, self.mappings["5GNR"]["USID"], "5GNR")
            nrcell = self.get_column_value(record, self.mappings["5GNR"]["cell"], "5GNR")
            if usid and nrcell:
                key = f"{usid}_{nrcell}"
                bbu_row = st.session_state.bbu_data[
                    (st.session_state.bbu_data[self.mappings["5GNR_BBU"]["USID"]].apply(lambda x: str(x)) == usid) &
                    (st.session_state.bbu_data[self.mappings["5GNR_BBU"]["NRCELL_NAME"]].apply(lambda x: str(x)) == nrcell)
                ]
                if not bbu_row.empty:
                    for name in self.mappings["5GNR_BBU"]["CONFIGURATION"]:
                        if name in bbu_row.iloc[0] and pd.notna(bbu_row.iloc[0][name]):
                            return str(bbu_row.iloc[0][name])
        return ""

    def perform_search(self):
        """Perform search based on user input"""
        if not st.session_state.search_value:
            st.error("Please enter a search value")
            return
        new_matched_records = []
        main_data = []
        lte_data = []
        nr_data = []
        def search_in_data(data, tech):
            mapping = self.mappings[tech][st.session_state.search_type]
            for _, record in data.iterrows():
                for col_name in mapping:
                    if col_name in record and pd.notna(record[col_name]) and \
                       st.session_state.search_value.lower() in str(record[col_name]).lower():
                        new_matched_records.append((tech, record.to_dict()))
                        row_data = {"Source": tech}
                        for key in self.mappings[tech]:
                            row_data[key] = self.get_column_value(record, self.mappings[tech][key], tech)
                        main_data.append(row_data)
                        if tech == "LTE":
                            lte_row = {
                                "Source": tech,
                                "Site": row_data["Site"],
                                "cell": row_data["cell"],
                                "CELLRANGE": row_data["CELLRANGE"],
                                "CRSGAIN": row_data["CRSGAIN"],
                                "QRXLEVMIN": row_data["QRXLEVMIN"],
                                "EARFCNDL": row_data["EARFCNDL"]
                            }
                            for col in st.session_state.lte_columns:
                                if col not in lte_row:
                                    lte_row[col] = self.get_column_value(record, [col], tech)
                            lte_data.append(lte_row)
                        elif tech == "5GNR":
                            nr_row = {
                                "Source": tech,
                                "USID": row_data["USID"],
                                "SITE": row_data["Site"],
                                "NRCELL_NAME": row_data["cell"],
                                "Digital Tilt": row_data["Digital Tilt"],
                                "Power": row_data["Power"],
                                "PCI": row_data["PCI"],
                                "ADMINISTRATIVESTATE": row_data["ADMINISTRATIVESTATE"],
                                "CELLBARRED": row_data["CELLBARRED"],
                                "CELLRESERVEDFOROPERATOR": row_data["CELLRESERVEDFOROPERATOR"],
                                "OPERATIONALSTATE": row_data["OPERATIONALSTATE"],
                                "CELLRANGE": row_data["CELLRANGE"],
                                "SSBFREQUENCY": row_data["SSBFREQUENCY"],
                                "CONFIGURATION": row_data["CONFIGURATION"]
                            }
                            for col in st.session_state.nr_columns:
                                if col not in nr_row:
                                    nr_row[col] = self.get_column_value(record, [col], tech)
                            nr_data.append(nr_row)
                        break
        if not st.session_state.lte_data.empty:
            search_in_data(st.session_state.lte_data, "LTE")
        if not st.session_state.nr_data.empty:
            search_in_data(st.session_state.nr_data, "5GNR")
        st.session_state.matched_records = new_matched_records
        if st.session_state.auto_generate:
            self.generate_vdt_data(lte_data, nr_data)
        self.update_status(f"Found {len(new_matched_records)} matching records")
        st.success(f"Found {len(new_matched_records)} matching records")

    def generate_vdt_data(self, lte_rows, nr_rows):
        """Generate VDT data"""
        lte_sites = sorted(set(row["Site"] for row in lte_rows if row["Site"]))
        nr_sites = sorted(set(row["SITE"] for row in nr_rows if row["SITE"]))
        max_len = max(len(lte_sites), len(nr_sites))
        vdt_rows = []
        for i in range(max_len):
            vdt_rows.append({
                "LTE Site": lte_sites[i] if i < len(lte_sites) else "",
                "NR Site": nr_sites[i] if i < len(nr_sites) else ""
            })
        st.session_state.vdt_data = pd.DataFrame(vdt_rows)
        self.update_status(f"Generated VDT data with {len(vdt_rows)} rows")

    def export_to_excel(self, df, filename):
        """Export DataFrame to Excel"""
        try:
            output = io.BytesIO()
            df.to_excel(output, index=False)
            output.seek(0)
            st.download_button(
                label=f"Download {filename}",
                data=output,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            self.update_status(f"Exported {len(df)} rows to {filename}")
        except Exception as e:
            logging.error(f"Error in export_to_excel: {str(e)}")
            st.error(f"Export failed: {str(e)}")

    def export_lte_to_excel(self):
        """Export LTE data to Excel"""
        lte_data = [
            {
                "Source": "LTE",
                "Site": self.get_column_value(record, self.mappings["LTE"]["Site"], "LTE"),
                "cell": self.get_column_value(record, self.mappings["LTE"]["cell"], "LTE"),
                "CELLRANGE": self.get_column_value(record, self.mappings["LTE"]["CELLRANGE"], "LTE"),
                "CRSGAIN": self.get_column_value(record, self.mappings["LTE"]["CRSGAIN"], "LTE"),
                "QRXLEVMIN": self.get_column_value(record, self.mappings["LTE"]["QRXLEVMIN"], "LTE"),
                "EARFCNDL": self.get_column_value(record, self.mappings["LTE"]["EARFCNDL"], "LTE")
            } for tech, record in st.session_state.matched_records if tech == "LTE"
        ]
        if not lte_data:
            st.error("No LTE data to export")
            return
        df = pd.DataFrame(lte_data, columns=st.session_state.lte_columns)
        self.export_to_excel(df, "LTE_Parameters.xlsx")

    def export_5g_to_excel(self):
        """Export 5G data to Excel"""
        nr_data = [
            {
                "Source": "5GNR",
                "USID": self.get_column_value(record, self.mappings["5GNR"]["USID"], "5GNR"),
                "SITE": self.get_column_value(record, self.mappings["5GNR"]["Site"], "5GNR"),
                "NRCELL_NAME": self.get_column_value(record, self.mappings["5GNR"]["cell"], "5GNR"),
                "Digital Tilt": self.get_column_value(record, self.mappings["5GNR"]["Digital Tilt"], "5GNR"),
                "Power": self.get_column_value(record, self.mappings["5GNR"]["Power"], "5GNR"),
                "PCI": self.get_column_value(record, self.mappings["5GNR"]["PCI"], "5GNR"),
                "ADMINISTRATIVESTATE": self.get_column_value(record, self.mappings["5GNR"]["ADMINISTRATIVESTATE"], "5GNR"),
                "CELLBARRED": self.get_column_value(record, self.mappings["5GNR"]["CELLBARRED"], "5GNR"),
                "CELLRESERVEDFOROPERATOR": self.get_column_value(record, self.mappings["5GNR"]["CELLRESERVEDFOROPERATOR"], "5GNR"),
                "OPERATIONALSTATE": self.get_column_value(record, self.mappings["5GNR"]["OPERATIONALSTATE"], "5GNR"),
                "CELLRANGE": self.get_column_value(record, self.mappings["5GNR"]["CELLRANGE"], "5GNR"),
                "SSBFREQUENCY": self.get_column_value(record, self.mappings["5GNR"]["SSBFREQUENCY"], "5GNR"),
                "CONFIGURATION": self.get_column_value(record, self.mappings["5GNR"]["CONFIGURATION"], "5GNR")
            } for tech, record in st.session_state.matched_records if tech == "5GNR"
        ]
        if not nr_data:
            st.error("No 5G data to export")
            return
        df = pd.DataFrame(nr_data, columns=st.session_state.nr_columns)
        self.export_to_excel(df, "5G_Parameters.xlsx")

    def export_lte_selected_to_excel(self, df):
        """Export selected LTE rows to Excel"""
        if df.empty:
            st.error("No rows selected for export")
            return
        self.export_to_excel(df, "LTE_Selected.xlsx")

    def export_5g_selected_to_excel(self, df):
        """Export selected 5G rows to Excel"""
        if df.empty:
            st.error("No rows selected for export")
            return
        self.export_to_excel(df, "5G_Selected.xlsx")

    def generate_lte_cr(self):
        """Generate LTE CR Excel"""
        lte_data = [
            {
                "Source": "LTE",
                "Site": self.get_column_value(record, self.mappings["LTE"]["Site"], "LTE"),
                "cell": self.get_column_value(record, self.mappings["LTE"]["cell"], "LTE"),
                "CELLRANGE": self.get_column_value(record, self.mappings["LTE"]["CELLRANGE"], "LTE"),
                "CRSGAIN": self.get_column_value(record, self.mappings["LTE"]["CRSGAIN"], "LTE"),
                "QRXLEVMIN": self.get_column_value(record, self.mappings["LTE"]["QRXLEVMIN"], "LTE"),
                "EARFCNDL": self.get_column_value(record, self.mappings["LTE"]["EARFCNDL"], "LTE"),
                "Electrical Tilt": self.get_column_value(record, self.mappings["LTE"]["Electrical Tilt"], "LTE")
            } for tech, record in st.session_state.matched_records if tech == "LTE"
        ]
        if not lte_data:
            st.error("No LTE rows selected")
            return
        cr_data = []
        for row in lte_data:
            site = row["Site"]
            cell = row["cell"]
            value = ""
            mo_class = f"EUtranCellFDD={cell}"
            if st.session_state.lte_cr_type == "cellRange":
                value = row["CELLRANGE"]
            elif st.session_state.lte_cr_type == "crsGain":
                value = row["CRSGAIN"]
            elif st.session_state.lte_cr_type == "Electrical Tilt":
                value = row["Electrical Tilt"]
            cr_data.append({
                "Site": site,
                "MO Class": mo_class,
                "Parameter": st.session_state.lte_cr_type,
                "Value": value,
                "CurrentValue": ""
            })
        if cr_data:
            df = pd.DataFrame(cr_data)
            self.export_to_excel(df, "LTE_CR.xlsx")
            self.update_status(f"Generated {len(cr_data)} LTE CRs for {st.session_state.lte_cr_type}")

    def generate_5g_cr(self):
        """Generate 5G CR Excel"""
        nr_data = [
            {
                "Source": "5GNR",
                "USID": self.get_column_value(record, self.mappings["5GNR"]["USID"], "5GNR"),
                "SITE": self.get_column_value(record, self.mappings["5GNR"]["Site"], "5GNR"),
                "NRCELL_NAME": self.get_column_value(record, self.mappings["5GNR"]["cell"], "5GNR"),
                "Digital Tilt": self.get_column_value(record, self.mappings["5GNR"]["Digital Tilt"], "5GNR"),
                "Power": self.get_column_value(record, self.mappings["5GNR"]["Power"], "5GNR"),
                "PCI": self.get_column_value(record, self.mappings["5GNR"]["PCI"], "5GNR"),
                "CELLRANGE": self.get_column_value(record, self.mappings["5GNR"]["CELLRANGE"], "5GNR")
            } for tech, record in st.session_state.matched_records if tech == "5GNR"
        ]
        if not nr_data:
            st.error("No 5G rows selected")
            return
        cr_data = []
        for row in nr_data:
            site = row["SITE"]
            cell = row["NRCELL_NAME"]
            value = ""
            mo_class = f"NRCellDU={cell}" if st.session_state.nr_cr_type == "cellRange" else f"NRSectorCarrier={cell},CommonBeamforming=1"
            if st.session_state.nr_cr_type == "digitalTilt":
                value = row["Digital Tilt"]
            elif st.session_state.nr_cr_type == "cellRange":
                value = row["CELLRANGE"]
            cr_data.append({
                "Site": site,
                "MO Class": mo_class,
                "Parameter": st.session_state.nr_cr_type,
                "Value": value,
                "CurrentValue": ""
            })
        if cr_data:
            df = pd.DataFrame(cr_data)
            self.export_to_excel(df, "5G_CR.xlsx")
            self.update_status(f"Generated {len(cr_data)} 5G CRs for {st.session_state.nr_cr_type}")

    def generate_vdt_report(self):
        """Generate VDT report Excel"""
        lte_sites = sorted(set(row["Site"] for row in [
            {
                "Site": self.get_column_value(record, self.mappings["LTE"]["Site"], "LTE")
            } for tech, record in st.session_state.matched_records if tech == "LTE"
        ] if row["Site"]))
        nr_sites = sorted(set(row["SITE"] for row in [
            {
                "SITE": self.get_column_value(record, self.mappings["5GNR"]["Site"], "5GNR")
            } for tech, record in st.session_state.matched_records if tech == "5GNR"
        ] if row["SITE"]))
        if not lte_sites and not nr_sites:
            st.error("No sites found in VDT tab. Please generate VDT data first.")
            return
        wb = openpyxl.Workbook()
        # LTE Sheet
        lte_data = [
            ["Type", "Value", "Remarks", "Buffer", ""],
            ["projectName", st.session_state.project_name, "", "0.01", ""],
            ["startTime", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Use exact format. Add \" ' \" before", "Default keep it 0.01. Can increase it to accommodate TA Plot", ""],
            ["endTime", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Use exact format. Add \" ' \" before", "", ""],
            ["Site Name List", "ZOOM (Default = 0 for automatic zoom, enter integer value for manual zoom\nEg. value: 5 will fetch plots up to 5 km diagonally from site in both directions.", 
             "Neighbour list (comma separated), if blank first tier neighbours will be considered", 
             "Carrier List(comma separated)", ""]
        ]
        for site in lte_sites:
            earfcns = sorted(set(
                self.get_column_value(record, self.mappings["LTE"]["EARFCNDL"], "LTE")
                for tech, record in st.session_state.matched_records
                if tech == "LTE" and self.get_column_value(record, self.mappings["LTE"]["Site"], "LTE") == site
                and self.get_column_value(record, self.mappings["LTE"]["EARFCNDL"], "LTE")
            ))
            lte_data.append([site, "0", "", ",".join(earfcns), ""])
        ws_lte = wb.create_sheet("LTE")
        for row in lte_data:
            ws_lte.append(row)
        # NR Sheet
        nr_data = [
            ["Type", "Value", "Remarks", "Buffer", "", ""],
            ["projectName", st.session_state.project_name, "", "0.01", "", ""],
            ["startTime", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Use exact format. Add \" ' \" before", "Default keep it 0.01. Can increase it to accommodate TA Plot", "", ""],
            ["endTime", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Use exact format. Add \" ' \" before", "", "", ""],
            ["Site Name List", "ZOOM (Default = 0 for automatic zoom, enter integer value for manual zoom\nEg. value: 5 will fetch plots up to 5 km diagonally from site in both directions.", 
             "Neighbour list (comma separated), if blank first tier neighbours will be considered", 
             "Carrier List(comma separated)", "", ""]
        ]
        for site in nr_sites:
            arfcns = sorted(set(
                self.get_column_value(record, self.mappings["5GNR"]["SSBFREQUENCY"], "5GNR")
                for tech, record in st.session_state.matched_records
                if tech == "5GNR" and self.get_column_value(record, self.mappings["5GNR"]["Site"], "5GNR") == site
                and self.get_column_value(record, self.mappings["5GNR"]["SSBFREQUENCY"], "5GNR")
            ))
            nr_data.append([site, "0", "", ",".join(arfcns), "", ""])
        ws_nr = wb.create_sheet("NR")
        for row in nr_data:
            ws_nr.append(row)
        wb.remove(wb["Sheet"])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        st.download_button(
            label="Download VDT Report",
            data=output,
            file_name="VDT_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        self.update_status(f"VDT report generated with {len(lte_sites)} LTE and {len(nr_sites)} NR sites")

    def add_point(self, name, lat, lon):
        """Add a point to the distance calculator"""
        try:
            if not name:
                name = f"Point {len(st.session_state.points) + 1}"
            if not lat or not lon:
                st.error("Please enter both latitude and longitude")
                return
            st.session_state.points.append((name, float(lat), float(lon)))
            self.update_status(f"Added point {name}")
        except Exception as e:
            logging.error(f"Error in add_point: {str(e)}")
            st.error(f"Invalid latitude or longitude: {str(e)}")

    def add_lat_long(self, lat_long, name):
        """Add point from lat,long string"""
        try:
            lat_long = lat_long.split(",")
            if len(lat_long) != 2:
                st.error("Invalid format. Use 'latitude,longitude'")
                return
            lat, lon = map(float, lat_long)
            self.add_point(name, lat, lon)
        except Exception as e:
            logging.error(f"Error in add_lat_long: {str(e)}")
            st.error(f"Invalid format: {str(e)}")

    def import_points_from_excel(self, file):
        """Import points from Excel"""
        try:
            df = pd.read_excel(file)
            new_points = []
            for _, row in df.iterrows():
                name = row.get("name") or row.get("Name") or row.get("point") or row.get("Point") or f"Point {len(st.session_state.points) + len(new_points) + 1}"
                lat = row.get("lat") or row.get("Latitude") or row.get("LAT")
                lon = row.get("lon") or row.get("Longitude") or row.get("LON") or row.get("long") or row.get("Long")
                if lat is not None and lon is not None:
                    new_points.append((name, float(lat), float(lon)))
            st.session_state.points.extend(new_points)
            self.update_status(f"Imported {len(new_points)} points from {file.name}")
        except Exception as e:
            logging.error(f"Error in import_points_from_excel: {str(e)}")
            st.error(f"Failed to import points: {str(e)}")

    def set_master_point(self, point):
        """Set master point"""
        st.session_state.master_point = point
        self.update_status(f"Set master point: {point[0]}")
        st.success(f"Set '{point[0]}' as the master point")

    def remove_points(self, selected):
        """Remove selected points"""
        if not selected:
            st.error("No points selected")
            return
        st.session_state.points = [p for p in st.session_state.points if p[0] not in selected]
        if st.session_state.master_point and st.session_state.master_point[0] in selected:
            st.session_state.master_point = None
        self.update_status(f"Removed {len(selected)} points")

    def clear_points(self):
        """Clear all points"""
        st.session_state.points = []
        st.session_state.master_point = None
        st.session_state.distance_results = ""
        self.update_status("Cleared all points")

    def calculate_distance(self, lat1, lon1, lat2, lon2):
        """Calculate distance between two points using Haversine formula"""
        try:
            R = 6371.0
            lat1_rad = radians(lat1)
            lon1_rad = radians(lon1)
            lat2_rad = radians(lat2)
            lon2_rad = radians(lon2)
            dlon = lon2_rad - lon1_rad
            dlat = lat2_rad - lat1_rad
            a = sin(dlat / 2)**2 + cos(lat1_rad) * cos(lat2_rad) * sin(dlon / 2)**2
            c = 2 * atan2(sqrt(a), sqrt(1 - a))
            return R * c
        except Exception as e:
            logging.error(f"Error in calculate_distance: {str(e)}")
            return 0

    def calculate_path_distances(self):
        """Calculate path distances between points"""
        if len(st.session_state.points) < 2:
            st.error("At least two points are required for distance calculation")
            return
        total_distance = 0
        results = ["Path Distances:"]
        for i in range(len(st.session_state.points) - 1):
            name1, lat1, lon1 = st.session_state.points[i]
            name2, lat2, lon2 = st.session_state.points[i + 1]
            distance = self.calculate_distance(lat1, lon1, lat2, lon2)
            total_distance += distance
            results.append(f"{name1} → {name2}: {distance:.3f} km")
        results.append(f"Total Distance: {total_distance:.3f} km")
        st.session_state.distance_results = "\n".join(results)
        self.update_status("Calculated path distances")

    def calculate_from_master(self):
        """Calculate distances from master point"""
        if not st.session_state.master_point:
            st.error("Please set a master point first")
            return
        if len(st.session_state.points) < 2:
            st.error("At least one additional point is required")
            return
        master_name, master_lat, master_lon = st.session_state.master_point
        total_distance = 0
        results = [f"Distances from Master Point ({master_name}):"]
        for name, lat, lon in st.session_state.points:
            if (name, lat, lon) == st.session_state.master_point:
                continue
            distance = self.calculate_distance(master_lat, master_lon, lat, lon)
            total_distance += distance
            results.append(f"{master_name} to {name}: {distance:.3f} km")
        results.append(f"Total Distance: {total_distance:.3f} km")
        st.session_state.distance_results = "\n".join(results)
        self.update_status("Calculated distances from master point")

    def add_lte_column(self):
        """Add a new column to LTE data"""
        data_sources = ["LTE", "5GNR", "5GNR_BBU"]
        source = st.selectbox("Select Data Source:", data_sources, key="lte_source_select")
        columns = []
        if source == "LTE" and not st.session_state.lte_data.empty:
            columns = sorted(st.session_state.lte_data.columns)
        elif source == "5GNR" and not st.session_state.nr_data.empty:
            columns = sorted(st.session_state.nr_data.columns)
        elif source == "5GNR_BBU" and not st.session_state.bbu_data.empty:
            columns = sorted(st.session_state.bbu_data.columns)
        if columns:
            column_name = st.selectbox("Select Column:", columns, key="lte_column_select")
            if column_name and column_name not in st.session_state.lte_columns:
                st.session_state.lte_columns.append(column_name)
                self.update_status(f"Added column {column_name} to LTE tab")
                st.success(f"Added column {column_name} to LTE tab")
            else:
                st.error(f"Column '{column_name}' already exists or is invalid")

    def add_5g_column(self):
        """Add a new column to 5G data"""
        data_sources = ["LTE", "5GNR", "5GNR_BBU"]
        source = st.selectbox("Select Data Source:", data_sources, key="nr_source_select")
        columns = []
        if source == "LTE" and not st.session_state.lte_data.empty:
            columns = sorted(st.session_state.lte_data.columns)
        elif source == "5GNR" and not st.session_state.nr_data.empty:
            columns = sorted(st.session_state.nr_data.columns)
        elif source == "5GNR_BBU" and not st.session_state.bbu_data.empty:
            columns = sorted(st.session_state.bbu_data.columns)
        if columns:
            column_name = st.selectbox("Select Column:", columns, key="nr_column_select")
            if column_name and column_name not in st.session_state.nr_columns:
                st.session_state.nr_columns.append(column_name)
                self.update_status(f"Added column {column_name} to 5G tab")
                st.success(f"Added column {column_name} to 5G tab")
            else:
                st.error(f"Column '{column_name}' already exists or is invalid")

    def generate_map_html(self):
        """Generate HTML with Google Maps showing all points with names and distances"""
        try:
            html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Network Points Map</title>
                <style>
                    body {{ margin: 0; padding: 0; }}
                    #map {{ height: 100vh; width: 100%; }}
                    #info-panel {{ position: absolute; top: 10px; left: 10px; background: white; padding: 10px; z-index: 1000; }}
                </style>
            </head>
            <body>
                <div id="info-panel"></div>
                <div id="map"></div>
                <script>
                    function initMap() {{
                        const map = new google.maps.Map(document.getElementById("map"), {{
                            zoom: {st.session_state.map_zoom},
                            center: {{ lat: {st.session_state.points[0][1] if st.session_state.points else 0}, lng: {st.session_state.points[0][2] if st.session_state.points else 0} }},
                            mapTypeId: '{st.session_state.map_type}'
                        }});
                        const infoWindow = new google.maps.InfoWindow();
                        const infoPanel = document.getElementById("info-panel");
                        const markers = [];
                        let infoContent = "<h3>Points and Distances</h3><ul>";
                        {self.generate_markers_js()}
                        {self.generate_paths_js()}
                        {self.generate_master_info_js()}
                        infoContent += "</ul>";
                        infoPanel.innerHTML = infoContent;
                    }}
                </script>
                <script async defer src="https://maps.googleapis.com/maps/api/js?key={self.api_key}&libraries=geometry&loading=async&callback=initMap"></script>
            </body>
            </html>
            """
            return html
        except Exception as e:
            logging.error(f"Error in generate_map_html: {str(e)}")
            return "<html><body>Error generating map</body></html>"

    def generate_markers_js(self):
        """Generate JavaScript code for markers"""
        try:
            js_code = ""
            for i, (name, lat, lon) in enumerate(st.session_state.points):
                js_code += f"""
                    const marker{i} = new google.maps.Marker({{
                        position: {{ lat: {lat}, lng: {lon} }},
                        map: map,
                        title: "{name}",
                        label: "{chr(65 + i)}"
                    }});
                    marker{i}.addListener('click', function() {{
                        infoWindow.setContent('<div><strong>{name}</strong><br>Lat: {lat:.6f}<br>Lon: {lon:.6f}</div>');
                        infoWindow.open(map, marker{i});
                    }});
                    markers.push(marker{i});
                    infoContent += "<li><strong>{name}</strong> (Lat: {lat:.6f}, Lon: {lon:.6f})</li>";
                """
            return js_code
        except Exception as e:
            logging.error(f"Error in generate_markers_js: {str(e)}")
            return ""

    def generate_paths_js(self):
        """Generate JavaScript code for paths and distances"""
        try:
            if len(st.session_state.points) < 2:
                return ""
            js_code = "const pathCoordinates = ["
            for _, lat, lon in st.session_state.points:
                js_code += f"{{lat: {lat}, lng: {lon}}},"
            js_code += "];"
            js_code += """
            const path = new google.maps.Polyline({
                path: pathCoordinates,
                geodesic: true,
                strokeColor: "#FF0000",
                strokeOpacity: 1.0,
                strokeWeight: 2
            });
            path.setMap(map);
            """
            for i in range(len(st.session_state.points) - 1):
                name1, lat1, lon1 = st.session_state.points[i]
                name2, lat2, lon2 = st.session_state.points[i + 1]
                distance = self.calculate_distance(lat1, lon1, lat2, lon2)
                mid_lat = (lat1 + lat2) / 2
                mid_lon = (lon1 + lon2) / 2
                js_code += f"""
                const distanceMarker{i} = new google.maps.Marker({{
                    position: {{ lat: {mid_lat}, lng: {mid_lon} }},
                    map: map,
                    icon: {{
                        path: google.maps.SymbolPath.CIRCLE,
                        scale: 0
                    }},
                    label: {{
                        text: "{distance:.2f} km",
                        color: "#0000FF",
                        fontSize: "12px",
                        fontWeight: "bold"
                    }}
                }});
                """
            return js_code
        except Exception as e:
            logging.error(f"Error in generate_paths_js: {str(e)}")
            return ""

    def generate_master_info_js(self):
        """Generate JavaScript code for master point info"""
        try:
            if not st.session_state.master_point:
                return ""
            master_name, master_lat, master_lon = st.session_state.master_point
            js_code = f"""
            const masterMarker = new google.maps.Marker({{
                position: {{ lat: {master_lat}, lng: {master_lon} }},
                map: map,
                title: "Master: {master_name}",
                label: "M",
                icon: {{
                    url: "http://maps.google.com/mapfiles/ms/icons/blue-dot.png"
                }}
            }});
            infoContent += "<h4>Distances from Master Point:</h4><ul>";
            """
            for name, lat, lon in st.session_state.points:
                if (name, lat, lon) == st.session_state.master_point:
                    continue
                distance = self.calculate_distance(master_lat, master_lon, lat, lon)
                js_code += f"""
                infoContent += "<li>{master_name} to {name}: {distance:.3f} km</li>";
                """
            js_code += "infoContent += \"</ul>\";"
            return js_code
        except Exception as e:
            logging.error(f"Error in generate_master_info_js: {str(e)}")
            return ""

    def show_map(self):
        """Show map with all points"""
        if not st.session_state.points:
            st.error("No points to show on map")
            return
        if not self.api_key:
            st.error("Google Maps API key is required for map visualization. Please set the environment variable GOOGLE_MAPS_API_KEY")
            return
        try:
            html_content = self.generate_map_html()
            with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.html') as f:
                f.write(html_content)
                temp_path = f.name
            st.markdown(f"[Open Map in Browser](file://{temp_path})")
            webbrowser.open(f"file://{temp_path}")
            self.update_status("Map opened in browser")
        except Exception as e:
            logging.error(f"Error in show_map: {str(e)}")
            st.error(f"Failed to load map: {str(e)}")

    def clear_results(self):
        """Clear all search results"""
        st.session_state.matched_records = []
        st.session_state.vdt_data = pd.DataFrame(columns=["LTE Site", "NR Site"])
        self.update_status("Cleared all results")
        st.success("Cleared all results")

def main():
    """Main function"""
    try:
        app = NetworkSearchApp()
    except Exception as e:
        error_msg = f"Critical Error: {str(e)}"
        logging.critical(error_msg)
        st.error(f"Application crashed: {str(e)}")
        with open("network_search_crash.log", "a") as f:
            f.write(f"{datetime.now()}: {error_msg}\n")

if __name__ == "__main__":
    main()