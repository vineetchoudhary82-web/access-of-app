import pandas as pd
import os
import streamlit as st
import logging
from datetime import datetime
import openpyxl
import math
from math import radians, sin, cos, sqrt, atan2
import io

# Configure logging
logging.basicConfig(filename='network_search.log', level=logging.ERROR,
                    format='%(asctime)s - %(levelname)s - %(message)s')

class NetworkSearchApp:
    def __init__(self):
        # Initialize session state variables
        if 'lte_data' not in st.session_state:
            st.session_state.lte_data = pd.DataFrame()
        if 'nr_data' not in st.session_state:
            st.session_state.nr_data = pd.DataFrame()
        if 'matched_records' not in st.session_state:
            st.session_state.matched_records = []
        if 'project_name' not in st.session_state:
            st.session_state.project_name = "ATT_STX_253"

        # Mappings for LTE and 5GNR data
        self.mappings = {
            "LTE": {"Site": ["MECONTEXT_ID", "SITE", "OSS_ENodeB"], "EARFCNDL": ["EARFCNDL"]},
            "5GNR": {"Site": ["CTS_COMMON_ID", "GNB_NAME", "GNODEB"], "SSBFREQUENCY": ["SSBFREQUENCY"]}
        }

        # Custom styling
        st.markdown(
            """
            <style>
            .stApp { background-color: #f0f4f8; }
            .stHeader { color: #1e90ff; font-size: 24px; font-weight: bold; }
            </style>
            """,
            unsafe_allow_html=True
        )

        self.create_widgets()

    def update_status(self, message):
        """Update status bar with timestamp"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        st.session_state.status = f"[{timestamp}] {message}"
        st.write(st.session_state.status)

    def create_widgets(self):
        st.markdown("<h1 class='stHeader'>Network Data Search Tool</h1>", unsafe_allow_html=True)

        # Data Loading Section
        with st.container():
            st.subheader("Data Loading")
            uploaded_files = st.file_uploader("Upload Multiple Files", accept_multiple_files=True, type=["xlsx", "xls"])
            if uploaded_files:
                st.session_state.file_paths = {file.name: file for file in uploaded_files}
                st.write("Uploaded Files:", ", ".join(st.session_state.file_paths.keys()))
            if st.button("Load Selected Data"):
                self.load_data(uploaded_files)

        # Search Section
        with st.container():
            st.subheader("Search")
            col1, col2, col3 = st.columns([1, 2, 1])
            with col1:
                search_type = st.selectbox("Search By:", ["Site"], key="search_type")
            with col2:
                search_value = st.text_input("Value:", key="search_value")
            with col3:
                if st.button("Search"):
                    self.perform_search(search_type, search_value)

        # VDT Tab
        with st.container():
            st.subheader("VDT Sheet")
            col1, col2 = st.columns([1, 1])
            with col1:
                project_name = st.selectbox("Project Name:", ["ATT_STX_253"], key="project_name")
            with col2:
                if st.button("Generate VDT Report"):
                    self.generate_vdt_report(project_name)

    def load_data(self, files):
        """Load data from uploaded Excel files"""
        try:
            new_lte_data = []
            new_nr_data = []
            for file in files:
                df = pd.read_excel(file)
                file_name = file.name.lower()
                if "lte" in file_name:
                    new_lte_data.append(df)
                elif "nr" in file_name:
                    new_nr_data.append(df)
            st.session_state.lte_data = pd.concat(new_lte_data, ignore_index=True) if new_lte_data else pd.DataFrame()
            st.session_state.nr_data = pd.concat(new_nr_data, ignore_index=True) if new_nr_data else pd.DataFrame()
            self.update_status(f"Loaded {len(files)} files")
            st.success("Data loading completed!")
        except Exception as e:
            logging.error(f"Error in load_data: {str(e)}")
            self.update_status(f"Error loading files: {str(e)}")
            st.error(f"Error loading files: {str(e)}")

    def get_column_value(self, record, possible_names, tech):
        """Retrieve the first valid column value from possible names"""
        for name in possible_names:
            if name in record and pd.notna(record[name]):
                return str(record[name])
        return ""

    def perform_search(self, search_type, search_value):
        """Perform search based on the selected type and value"""
        if not search_value:
            st.error("Please enter a search value")
            return
        new_matched_records = []
        def search_in_data(data, tech):
            mapping = self.mappings[tech][search_type]
            for _, record in data.iterrows():
                for col_name in mapping:
                    if col_name in record and pd.notna(record[col_name]) and search_value.lower() in str(record[col_name]).lower():
                        new_matched_records.append((tech, record.to_dict()))
                        break
        if not st.session_state.lte_data.empty:
            search_in_data(st.session_state.lte_data, "LTE")
        if not st.session_state.nr_data.empty:
            search_in_data(st.session_state.nr_data, "5GNR")
        st.session_state.matched_records = new_matched_records
        self.update_status(f"Found {len(new_matched_records)} matching records")
        st.success(f"Found {len(new_matched_records)} matching records")

    def generate_vdt_report(self, project_name):
        """Generate and download the VDT report as an Excel file"""
        try:
            lte_sites = sorted(set(self.get_column_value(record, self.mappings["LTE"]["Site"], "LTE")
                                 for tech, record in st.session_state.matched_records
                                 if tech == "LTE" and self.get_column_value(record, self.mappings["LTE"]["Site"], "LTE")))
            nr_sites = sorted(set(self.get_column_value(record, self.mappings["5GNR"]["Site"], "5GNR")
                                for tech, record in st.session_state.matched_records
                                if tech == "5GNR" and self.get_column_value(record, self.mappings["5GNR"]["Site"], "5GNR")))
            if not lte_sites and not nr_sites:
                st.error("No sites found. Please perform a search first.")
                return

            wb = openpyxl.Workbook()
            # LTE Sheet
            lte_data = [
                ["Type", "Value", "Remarks", "Buffer", ""],
                ["projectName", project_name, "", "0.01", ""],
                ["startTime", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Use exact format. Add \" ' \" before", "Default keep it 0.01. Can increase it to accommodate TA Plot", ""],
                ["endTime", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Use exact format. Add \" ' \" before", "", ""],
                ["Site Name List", "ZOOM (Default = 0 for automatic zoom, enter integer value for manual zoom\nEg. value: 5 will fetch plots up to 5 km diagonally from site in both directions.)",
                 "Neighbour list (comma separated), if blank first tier neighbours will be considered", "Carrier List(comma separated)", ""]
            ]
            for site in lte_sites:
                earfcns = sorted(set(self.get_column_value(record, self.mappings["LTE"]["EARFCNDL"], "LTE")
                                   for tech, record in st.session_state.matched_records
                                   if tech == "LTE" and self.get_column_value(record, self.mappings["LTE"]["Site"], "LTE") == site
                                   and self.get_column_value(record, self.mappings["LTE"]["EARFCNDL"], "LTE")))
                lte_data.append([site, "0", "", ",".join(earfcns), ""])
            ws_lte = wb.create_sheet("LTE")
            for row in lte_data:
                ws_lte.append(row)

            # NR Sheet (Fixed at line 814)
            nr_data = [
                ["Type", "Value", "Remarks", "Buffer", "", ""],
                ["projectName", st.session_state.project_name, "", "0.01", "", ""],
                ["startTime", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Use exact format. Add \" ' \" before", "Default keep it 0.01. Can increase it to accommodate TA Plot", "", ""],
                ["endTime", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Use exact format. Add \" ' \" before", "", "", ""],
                ["Site Name List", "ZOOM (Default = 0 for automatic zoom, enter integer value for manual zoom\nEg. value: 5 will fetch plots up to 5 km diagonally from site in both directions.)",
                 "Neighbour list (comma separated), if blank first tier neighbours will be considered",
                 "Carrier List(comma separated)", "", ""]
            ]
            for site in nr_sites:
                arfcns = sorted(set(self.get_column_value(record, self.mappings["5GNR"]["SSBFREQUENCY"], "5GNR")
                                  for tech, record in st.session_state.matched_records
                                  if tech == "5GNR" and self.get_column_value(record, self.mappings["5GNR"]["Site"], "5GNR") == site
                                  and self.get_column_value(record, self.mappings["5GNR"]["SSBFREQUENCY"], "5GNR")))
                nr_data.append([site, "0", "", ",".join(arfcns), "", ""])
            ws_nr = wb.create_sheet("NR")
            for row in nr_data:
                ws_nr.append(row)

            wb.remove(wb["Sheet"])
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            st.download_button(label="Download VDT Report", data=output, file_name="VDT_Report.xlsx",
                             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.update_status(f"VDT report generated with {len(lte_sites)} LTE and {len(nr_sites)} NR sites")
        except Exception as e:
            logging.error(f"Error in generate_vdt_report: {str(e)}")
            self.update_status(f"Error generating VDT report: {str(e)}")
            st.error(f"Error generating VDT report: {str(e)}")

def main():
    """Main function with exception handling"""
    try:
        app = NetworkSearchApp()
    except Exception as e:
        logging.critical(f"Critical Error: {str(e)}\n{traceback.format_exc()}")
        st.error(f"Application crashed: {str(e)}")

if __name__ == "__main__":
    main()